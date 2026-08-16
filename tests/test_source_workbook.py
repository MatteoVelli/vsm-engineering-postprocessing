from pathlib import Path

import pytest
from openpyxl import Workbook

from vsm_postprocessing.errors import DataValidationError
from vsm_postprocessing.importer import ImportOptions, inspect_data_file
from conftest import CAIMAN_REFERENCE_DESCRIPTION, CAIMAN_REFERENCE_XLSX, require_private_reference_file


OLD_HYBRID_WORKBOOK = (
    Path(__file__).resolve().parents[1] / "reference_files" / "Robo_Sprayer_Electrification_Tamplate_Hybrid.xlsx"
)
CORRECTED_HYBRID_WORKBOOK = (
    Path(__file__).resolve().parents[1] / "reference_files" / "Robo_Sprayer_Electrification_Tamplate_Hybrid_02.xlsx"
)


def test_supplied_source_workbook_acceptance_criteria() -> None:
    source_workbook = require_private_reference_file(CAIMAN_REFERENCE_XLSX, CAIMAN_REFERENCE_DESCRIPTION)
    result = inspect_data_file(source_workbook)
    quality = result.quality

    assert quality.is_valid
    assert quality.sheet_name == "Sprayer_Caiman_SP_9300Kg_Hybrid"
    assert quality.header_row == 3
    assert quality.unit_row == 4
    assert quality.data_start_row == 5
    assert quality.data_end_row == 1870
    assert quality.sample_count == 1866
    assert quality.channel_count == 70
    assert quality.raw_channel_count == 45
    assert quality.math_channel_count == 25
    assert quality.time_channel_name == "Track_Time"
    assert quality.time_unit == "s"
    assert quality.time_start == 0.0
    assert quality.time_end == 1865.0
    assert quality.nominal_time_step == 1.0
    assert quality.time_is_strictly_increasing
    assert quality.duplicate_timestamp_count == 0
    assert quality.missing_cell_count == 0
    assert quality.invalid_numeric_cell_count == 0
    assert quality.non_finite_cell_count == 0

    channel_ids = [channel.channel_id for channel in result.channels]
    assert len(channel_ids) == len(set(channel_ids))


@pytest.mark.skipif(
    not OLD_HYBRID_WORKBOOK.exists() or not CORRECTED_HYBRID_WORKBOOK.exists(),
    reason="RoboSprayer Hybrid workbooks are not present",
)
def test_corrected_hybrid_workbook_generator_torque_mapping_supersedes_old_source() -> None:
    from openpyxl import load_workbook

    old_book = load_workbook(OLD_HYBRID_WORKBOOK, read_only=True, data_only=False)
    corrected_book = load_workbook(CORRECTED_HYBRID_WORKBOOK, read_only=True, data_only=False)
    try:
        old_mapping = old_book["Sheet1"]
        corrected_mapping = corrected_book["Rename From VSM to Astauto"]
        old_data = old_book["Hybrid_1C2G_30-60kph (2)"]
        corrected_data = corrected_book["Hybrid_1C2G_30-60kph (2)"]

        assert old_mapping.cell(226, 2).value == "Generator Torque_1"
        assert corrected_mapping.cell(226, 2).value == "Engine_AuxiliaryTorque_1"
        assert corrected_mapping.cell(226, 3).value == "ICE Generator Torque "
        assert corrected_mapping.cell(226, 4).value == "AVL"
        assert corrected_mapping.cell(226, 5).value == "√"

        assert old_data["HQ5"].value == corrected_data["HQ5"].value == "=BX5*HP5/9548.8"
    finally:
        old_book.close()
        corrected_book.close()


def test_nonmonotonic_time_fails_strict_validation(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Track_Time", "Speed"])
    sheet.append(["s", "kph"])
    sheet.append([0, 0])
    sheet.append([1, 10])
    sheet.append([1, 20])
    path = tmp_path / "nonmonotonic.xlsx"
    workbook.save(path)

    with pytest.raises(DataValidationError, match="not strictly increasing"):
        inspect_data_file(path, ImportOptions(header_row=1, unit_row=2, data_start_row=3))


def test_duplicate_display_names_still_receive_unique_ids(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Track_Time", "Power", "Power"])
    sheet.append(["s", "kW", "kW"])
    sheet.append([0, 1, 2])
    sheet.append([1, 3, 4])
    path = tmp_path / "duplicates.xlsx"
    workbook.save(path)

    result = inspect_data_file(path, ImportOptions(header_row=1, unit_row=2, data_start_row=3))
    ids = [channel.channel_id for channel in result.channels]

    assert len(ids) == len(set(ids))
    assert result.channels[1].source_name == result.channels[2].source_name == "Power"


def test_csv_import_with_units(tmp_path: Path) -> None:
    path = tmp_path / "data.csv"
    path.write_text("Track_Time;Speed\ns;kph\n0;0\n1;12.5\n2;20\n", encoding="utf-8")

    result = inspect_data_file(path)

    assert result.quality.is_valid
    assert result.quality.sample_count == 3
    assert result.quality.channel_count == 2
    assert result.quality.raw_channel_count == 2
    assert result.quality.math_channel_count == 0
    assert result.quality.nominal_time_step == 1.0
