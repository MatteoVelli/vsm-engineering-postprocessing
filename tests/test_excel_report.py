from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from vsm_postprocessing.errors import ConfigurationError
from vsm_postprocessing.excel_report_engine import generate_excel_report, load_excel_report_config
from vsm_postprocessing.importer import ImportOptions


def _write_config(path: Path, *, channels: str = "  - time__col_001\n", output: str = "report.xlsx", bottom: str = "    - max\n") -> None:
    path.write_text(
        f"""version: 1
report:
  title: Test Report
  report_sheet: Report
  metadata_sheet: Metadata
channels:
{channels}statistics:
  top_rms: []
  kpis: []
  bottom_operations:
{bottom}plots:
  include: []
  columns: 2
  width_px: 600
  height_px: 300
output:
  filename: {output}
  keep_plot_assets: true
""",
        encoding="utf-8",
    )


def test_excel_report_config_loads(tmp_path: Path) -> None:
    path = tmp_path / "report.yaml"
    _write_config(path)
    config = load_excel_report_config(path)
    assert config.version == 1
    assert config.report_sheet == "Report"
    assert config.metadata_sheet == "Metadata"
    assert config.channel_ids == ("time__col_001",)
    assert config.bottom_operations == ("max",)
    assert config.output_filename == "report.xlsx"


def test_duplicate_report_channels_are_rejected(tmp_path: Path) -> None:
    path = tmp_path / "report.yaml"
    _write_config(path, channels="  - time__col_001\n  - time__col_001\n")
    with pytest.raises(ConfigurationError, match="duplicate channel IDs"):
        load_excel_report_config(path)


def test_invalid_bottom_operation_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "report.yaml"
    _write_config(path, bottom="    - median\n")
    with pytest.raises(ConfigurationError, match="unsupported operations"):
        load_excel_report_config(path)


def test_output_must_be_plain_xlsx_filename(tmp_path: Path) -> None:
    path = tmp_path / "report.yaml"
    _write_config(path, output="nested/report.xlsx")
    with pytest.raises(ConfigurationError, match="plain .xlsx filename"):
        load_excel_report_config(path)


def test_sergio_reference_layout_config_loads(tmp_path: Path) -> None:
    path = tmp_path / "report.yaml"
    path.write_text(
        """version: 1
report:
  title: Test Report
  report_sheet: Report
  metadata_sheet: Metadata
channels:
  - time__col_001
statistics:
  top_rms: []
  kpis: []
  bottom_operations:
    - max
  bottom_summary: []
layout:
  profile: sergio_reference
  plot_placement: kpi_panel
  blank_separator_columns: 1
  channel_width: 12
  header_row_height: 106
  unit_row_height: 16
plots:
  include: []
  columns: 2
  width_px: 600
  height_px: 300
output:
  filename: report.xlsx
  keep_plot_assets: true
""",
        encoding="utf-8",
    )
    config = load_excel_report_config(path)
    assert config.layout_profile == "sergio_reference"
    assert config.plot_placement == "kpi_panel"
    assert config.header_row_height == 106
    assert config.unit_row_height == 16


def test_invalid_layout_profile_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "report.yaml"
    path.write_text(
        """version: 1
report:
  title: Test Report
channels:
  - time__col_001
statistics:
  top_rms: []
  kpis: []
  bottom_operations:
    - max
plots:
  include: []
layout:
  profile: copied_by_hand
output:
  filename: report.xlsx
""",
        encoding="utf-8",
    )
    with pytest.raises(ConfigurationError, match="layout.profile"):
        load_excel_report_config(path)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = PROJECT_ROOT / "reference_files" / (
    "Sprayer_Caiman_SP_9300Kg_Hybrid_Gen80kW_30kph_74Ht_4000KgAQ_57-4pcSOC_5-80_1C2G_02.xlsx"
)
REPORT_CONFIG = PROJECT_ROOT / "config" / "excel_report_example.yaml"
MATH_CONFIG = PROJECT_ROOT / "config" / "math_channels_example.yaml"
STATISTICS_CONFIG = PROJECT_ROOT / "config" / "statistics_excel_report.yaml"
PLOTTING_CONFIG = PROJECT_ROOT / "config" / "plotting_example.yaml"


@pytest.mark.skipif(not SOURCE_WORKBOOK.exists(), reason="Client source workbook is not present")
def test_supplied_source_workbook_excel_report_acceptance(tmp_path: Path) -> None:
    result = generate_excel_report(
        SOURCE_WORKBOOK,
        REPORT_CONFIG,
        STATISTICS_CONFIG,
        PLOTTING_CONFIG,
        tmp_path / "excel_report",
        ImportOptions(strict=True),
        math_config_file=MATH_CONFIG,
    )

    assert result.sample_count == 1866
    assert result.channel_count == 21
    assert result.statistic_count == 53
    assert result.plot_count == 24
    assert result.report_path.exists()
    assert result.manifest_path.exists()
    assert result.summary_path.exists()

    workbook = load_workbook(result.report_path, data_only=True)
    try:
        assert workbook.sheetnames == ["Report", "Metadata"]
        report = workbook["Report"]
        metadata = workbook["Metadata"]
        assert report.freeze_panes == "B6"
        assert report["A1"].value is None
        assert report["A3"].value == "Track_Time"
        assert report["A4"].value == "s"
        assert report["A5"].value == pytest.approx(0.0)
        assert report["A1870"].value == pytest.approx(1865.0)
        assert report["F1"].value == "Battery Power RMS"
        assert report["F2"].value == pytest.approx(64.9480711679)
        assert report["H1"].value == "Battery Heatflow RMS"
        assert report["H2"].value == pytest.approx(6.4230535881)
        assert report["C1871"].value == pytest.approx(30.0717)
        assert report["F1871"].value == pytest.approx(23.5799)
        assert report["F1872"].value == pytest.approx(-149.18)
        assert report["H1871"].value == pytest.approx(14.7541)
        assert report["J1871"].value == pytest.approx(23.9383)
        assert report["W3"].value == "Time [min]"
        assert report["W4"].value == pytest.approx(1865.0 / 60.0)
        assert len(report._images) == 6
        assert metadata["A1"].value == "VSM REPORT METADATA"
    finally:
        workbook.close()
