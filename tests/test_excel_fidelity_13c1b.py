from __future__ import annotations

import re
import json
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from vsm_postprocessing.excel_report_engine import load_excel_report_config


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REPORT = PROJECT_ROOT / "outputs" / "end_to_end_sergio_duty_cycle" / "07_excel_report" / "vsm_engineering_report.xlsx"
CONFIG = PROJECT_ROOT / "config" / "excel_report_duty_cycle.yaml"


@pytest.fixture(scope="module")
def workbook():
    if not REPORT.exists():
        pytest.skip("Generated Sergio duty-cycle workbook is not present")
    wb = load_workbook(REPORT, data_only=True)
    try:
        yield wb
    finally:
        wb.close()


def _chart_xml_roots() -> list[ET.Element]:
    if not REPORT.exists():
        pytest.skip("Generated Sergio duty-cycle workbook is not present")
    ns = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"
    with zipfile.ZipFile(REPORT) as archive:
        names = sorted(
            (name for name in archive.namelist() if name.startswith("xl/charts/chart") and name.endswith(".xml")),
            key=lambda value: int(re.search(r"chart(\d+)\.xml", value).group(1)),
        )
        roots = [ET.fromstring(archive.read(name)) for name in names]
    assert len(roots) == 18
    assert ns
    return roots


def test_13c1b_native_secondary_axes_and_bottom_x_axis() -> None:
    ns = {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}
    roots = _chart_xml_roots()
    secondary = 0
    for root in roots:
        groups = root.findall(".//c:scatterChart", ns)
        val_axes = root.findall(".//c:valAx", ns)
        bottom_axes = [axis for axis in val_axes if axis.find("c:axPos", ns) is not None and axis.find("c:axPos", ns).get("val") == "b"]
        right_axes = [axis for axis in val_axes if axis.find("c:axPos", ns) is not None and axis.find("c:axPos", ns).get("val") == "r"]
        assert bottom_axes
        if len(groups) > 1:
            secondary += 1
            assert right_axes
            assert any(
                axis.find("c:crosses", ns) is not None and axis.find("c:crosses", ns).get("val") == "max"
                for axis in right_axes
            )
    assert secondary == 12


def test_13c1b_chart_layout_is_reference_sized_and_outside_data_table(workbook) -> None:
    config = load_excel_report_config(CONFIG)
    ws = workbook["Report"]
    assert len(config.chart_layout) == 18
    assert len(ws._charts) == 18
    for chart in ws._charts:
        anchor = chart.anchor._from
        start_col = anchor.col + 1
        assert start_col > column_index_from_string("BR")
        assert anchor.row + 1 >= 6
        assert chart.width >= 15.0
        assert chart.height >= 7.5
    anchors = {f"{get_column_letter(chart.anchor._from.col + 1)}{chart.anchor._from.row + 1}" for chart in ws._charts}
    assert {"BT6", "CA6", "CH6", "CO6", "CW6", "DE6", "DS6", "EA7"} <= anchors


def test_13c1b_kpi_count_freeze_and_rms_merges(workbook) -> None:
    config = load_excel_report_config(CONFIG)
    ws = workbook["Report"]
    assert len(config.kpi_statistic_ids) == 34
    assert ws.freeze_panes == "B6"
    assert {str(item) for item in ws.merged_cells.ranges} >= {"L1:N1", "P1:S1"}
    assert ws["BT3"].value == "Time [min]"
    assert ws["BV3"].value == "Max Speed [kph]"
    assert ws["BX3"].value == "Max Engine Speed [rpm]"
    assert ws["BY3"].value == "Max Engine Torque [Nm]"
    assert ws["CA3"].value == "Engine Energy Delivered [kWh]"
    assert ws["CB3"].value == "Max Total Generator Power [kW]"
    assert ws["DA3"].value == "Max Generator Power 1 [kW]"


def test_13c1b_bottom_summary_coverage(workbook) -> None:
    ws = workbook["Report"]
    required = ["F", "J", "K", "Q", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AW", "BI", "BJ", "BN", "BP"]
    assert sum(1 for col in range(1, 71) if ws.cell(17423, col).value not in (None, "")) >= 48
    for column in required:
        assert ws[f"{column}17423"].value not in (None, "")
    assert ws["M17424"].value == pytest.approx(-149.18)


def test_13c1b_client_safe_provenance(workbook) -> None:
    metadata = workbook["Metadata"]
    values = [str(cell.value) for row in metadata.iter_rows() for cell in row if cell.value is not None]
    assert any(value == "Original VSM source workbook" for value in values)
    assert any("Sprayer_Caiman_SP_9300Kg_Hybrid" in value for value in values)
    assert any(value == "External profile/reference workbook" for value in values)
    assert any(value == "Sprayer_Caiman_SP_9300Kg_Electrification_03.xlsx" for value in values)
    assert not any("C:\\Users" in value or "Desktop\\Agro" in value for value in values)


def test_13c1c_excel_summary_and_manifest_use_explicit_chart_counts() -> None:
    manifest_path = REPORT.with_name("excel_report_manifest.json")
    summary_path = REPORT.with_name("excel_report_summary.txt")
    if not manifest_path.exists() or not summary_path.exists():
        pytest.skip("Generated Sergio duty-cycle Excel summary artifacts are not present")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    summary = summary_path.read_text(encoding="utf-8")
    assert manifest["native_excel_chart_count"] == 18
    assert manifest["configured_plot_count"] == 24
    assert manifest["plot_series_count"] == 45
    assert manifest["embedded_plot_image_count"] == 0
    assert "Native Excel charts embedded: 18" in summary
    assert "Configured plots rendered: 24" in summary
    assert "Plots embedded: 24" not in summary
