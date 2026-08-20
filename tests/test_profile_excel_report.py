from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from vsm_postprocessing.excel_report_engine import generate_profile_excel_report
from vsm_postprocessing.importer import ImportOptions

from conftest import (
    ROBOSPRAYER_LATEST_ELECTRIC_CSV,
    ROBOSPRAYER_LATEST_ELECTRIC_DESCRIPTION,
    ROBOSPRAYER_LATEST_HYBRID_CSV,
    ROBOSPRAYER_LATEST_HYBRID_DESCRIPTION,
    ROBOSPRAYER_REFERENCE_CSV,
    ROBOSPRAYER_REFERENCE_DESCRIPTION,
    require_private_reference_file,
)
ELECTRIC_PROFILE = Path("config/report_profiles/robosprayer_electric.yaml")
HYBRID_PROFILE = Path("config/report_profiles/robosprayer_hybrid.yaml")
DATA_START_ROW = 5


def _robosprayer_csv() -> Path:
    return require_private_reference_file(ROBOSPRAYER_REFERENCE_CSV, ROBOSPRAYER_REFERENCE_DESCRIPTION)


def _latest_electric_csv() -> Path:
    return require_private_reference_file(ROBOSPRAYER_LATEST_ELECTRIC_CSV, ROBOSPRAYER_LATEST_ELECTRIC_DESCRIPTION)


def _latest_hybrid_csv() -> Path:
    return require_private_reference_file(ROBOSPRAYER_LATEST_HYBRID_CSV, ROBOSPRAYER_LATEST_HYBRID_DESCRIPTION)


@pytest.fixture(scope="module")
def electric_report(tmp_path_factory: pytest.TempPathFactory):
    return generate_profile_excel_report(
        _robosprayer_csv(),
        ELECTRIC_PROFILE,
        tmp_path_factory.mktemp("profile_excel_electric"),
        ImportOptions(),
    )


@pytest.fixture(scope="module")
def electric_workbook(electric_report):
    return load_workbook(electric_report.report_path, data_only=True)


def test_profile_excel_report_generates_reopenable_electric_workbook(electric_report, electric_workbook) -> None:
    assert electric_report.report_path.exists()
    assert electric_report.report_path.name == "RoboSprayer_Electric_Engineering_Report.xlsx"
    assert electric_report.sample_count == 3853
    assert electric_report.source_raw_channel_count == 607
    assert electric_report.report_channel_count == 317
    assert electric_report.vsm_count == 182
    assert electric_report.avl_count == 110
    assert electric_report.math_count == 25
    assert electric_report.statistic_count == 27
    assert electric_report.kpi_count == 9
    assert electric_report.plot_count == 14
    assert electric_workbook.sheetnames == [
        "RoboSprayer Electric",
        "Rename From VSM to Astauto",
        "Metadata",
    ]
    assert [sheet.title for sheet in electric_workbook.worksheets if sheet.sheet_state == "visible"] == [
        "RoboSprayer Electric",
        "Rename From VSM to Astauto",
    ]
    assert [sheet.title for sheet in electric_workbook.worksheets if sheet.sheet_state == "hidden"] == ["Metadata"]


def test_profile_excel_report_exports_semantic_raw_and_math_channels_with_dynamic_rows(
    electric_report,
    electric_workbook,
) -> None:
    sheet = electric_workbook["RoboSprayer Electric"]
    mapping = electric_workbook["Rename From VSM to Astauto"]
    channel_types = [mapping.cell(row, 3).value for row in range(3, 320)]

    assert channel_types.count("VSM") == 182
    assert channel_types.count("AVL") == 110
    assert channel_types.count("MATH") == 25
    assert sheet.cell(3, 1).value == "Time"
    assert sheet.cell(3, 2).value == "Time"
    assert sheet.cell(4, 2).value == "min"
    assert not any("__col_" in str(sheet.cell(3, col).value) for col in range(1, electric_report.report_channel_count + 1))
    assert electric_report.report_channel_count == 317
    assert get_column_letter(electric_report.report_channel_count) == "LE"
    assert sheet.max_column > electric_report.report_channel_count
    assert sheet.cell(DATA_START_ROW + 3853 - 1, 1).value == pytest.approx(3852.0)
    assert sheet.cell(DATA_START_ROW + 3853, 1).value == "MAX"

    by_semantic = {channel.channel_id: index + 1 for index, channel in enumerate(electric_report.report_channels)}
    assert by_semantic["time_minutes"] == 2
    assert by_semantic["distance_km"] == 7
    assert by_semantic["total_edu_elect_power"] == 28
    assert by_semantic["battery_power_squared"] == 30
    assert by_semantic["battery_heatflow_squared"] == 32
    assert by_semantic["total_edu_mech_power"] == 65
    assert by_semantic["agrochemical_discharge"] == 74
    data_end_row = DATA_START_ROW + electric_report.sample_count - 1
    assert sheet.cell(DATA_START_ROW, by_semantic["track_time"]).value == pytest.approx(0.0)
    assert sheet.cell(data_end_row, by_semantic["time_minutes"]).value == pytest.approx(64.2)
    assert sheet.cell(data_end_row, by_semantic["distance_km"]).value == pytest.approx(11.9996)


def test_profile_excel_report_statistics_kpis_and_correct_rms_values(electric_report, electric_workbook) -> None:
    sheet = electric_workbook["RoboSprayer Electric"]
    by_semantic = {channel.channel_id: index + 1 for index, channel in enumerate(electric_report.report_channels)}
    right_summary = {sheet.cell(3, col).value: sheet.cell(4, col).value for col in range(319, 355)}

    assert "AB1:AD1" in {str(item) for item in sheet.merged_cells.ranges}
    assert "AF1:AI1" in {str(item) for item in sheet.merged_cells.ranges}
    assert sheet.cell(2, 30).value == pytest.approx(28.716636645770492)
    assert sheet.cell(2, 30).value != pytest.approx(13.506611297860566)
    assert sheet.cell(2, 32).value == pytest.approx(2.840107066457429)
    assert sheet.cell(2, 32).value != pytest.approx(1.3358187681981721)
    assert sheet.cell(3858, by_semantic["chassis_speed"]).value == pytest.approx(11.9858)
    assert sheet.cell(3859, by_semantic["electricsystem_battery_power"]).value == pytest.approx(-29.854)
    assert sheet.cell(3860, by_semantic["electricsystem_battery_soc"]).value == pytest.approx(12.6927)
    assert sheet.cell(3861, by_semantic["electricsystem_battery_energy"]).value == pytest.approx(40.0)
    assert sheet.cell(3858, by_semantic["agrochemical_discharge"]).value == pytest.approx(0.0)
    assert right_summary["Battery Capacity Used [kWh]"] == pytest.approx(33.65367)
    assert right_summary["Battery Energy Consumption [Wh/Km]"] == pytest.approx(2804.565985532851)
    assert right_summary["100% Battery Capacity [kWh]"] == pytest.approx(50.0)
    assert right_summary["Range for 85% Battery [Km]"] == pytest.approx(15.153859891060916)


def test_profile_excel_report_includes_plots_metadata_and_template_comparison(
    electric_report,
    electric_workbook,
) -> None:
    report_sheet = electric_workbook["RoboSprayer Electric"]
    metadata_sheet = electric_workbook["Metadata"]

    assert len(report_sheet._images) == 14
    anchors = [(image.anchor._from.col + 1, image.anchor._from.row + 1) for image in report_sheet._images]
    assert anchors[:6] == [(319, 7), (325, 7), (331, 7), (337, 7), (343, 7), (349, 7)]
    assert anchors[6:12] == [(319, 24), (325, 24), (331, 24), (337, 24), (343, 24), (349, 24)]
    assert anchors[12:] == [(319, 41), (325, 41)]
    assert report_sheet.cell(6, 319).value == "Speed Vs Distance"
    assert report_sheet.cell(6, 331).value == "Road Profile"
    assert report_sheet.cell(40, 319).value == "Agrochemical Discharge and Battery SOC Vs Time"
    assert report_sheet.cell(6, 331).fill.fgColor.rgb == "001F4E78"
    assert report_sheet.cell(6, 331).font.color.rgb == "00FFFFFF"
    assert report_sheet.cell(6, 331).font.bold is True
    assert report_sheet._images[0].anchor.ext.cx == 405 * 9525
    assert report_sheet._images[0].anchor.ext.cy == 234 * 9525

    metadata = {metadata_sheet.cell(row, 1).value: metadata_sheet.cell(row, 2).value for row in range(2, 45)}
    assert metadata["Source sample count"] == 3853
    assert metadata["Workbook data start row"] == DATA_START_ROW
    assert metadata["Workbook data end row"] == DATA_START_ROW + 3853 - 1
    assert metadata["Exported report channels"] == 317
    assert metadata["Rendered plots"] == 14
    assert metadata["Resolved raw profile channels"] == (
        metadata["Exported VSM raw channels"] + metadata["Exported AVL raw channels"]
    )
    assert metadata["Report title"] == "RoboSprayer Electric"

    visible_text = []
    visible_formulas = []
    for sheet in [worksheet for worksheet in electric_workbook.worksheets if worksheet.sheet_state == "visible"]:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    visible_text.append(cell.value)
                    if cell.value.startswith("="):
                        visible_formulas.append((sheet.title, cell.coordinate, cell.value))
    assert not any("C:\\Users\\" in value or "Desktop\\Agro Project" in value for value in visible_text)
    assert visible_formulas == []
    assert "Template Comparison" not in electric_workbook.sheetnames
    assert "INTENTIONAL CORRECTION" in {row["status"] for row in electric_report.template_comparison}


def test_profile_excel_report_hybrid_dry_run_remains_profile_generic(tmp_path: Path) -> None:
    result = generate_profile_excel_report(
        _robosprayer_csv(),
        HYBRID_PROFILE,
        tmp_path / "profile_excel_hybrid",
        ImportOptions(),
    )
    workbook = load_workbook(result.report_path, data_only=True)

    assert result.report_path.name == "RoboSprayer_Hybrid_Engineering_Report.xlsx"
    assert workbook.sheetnames[0] == "RoboSprayer Hybrid"
    assert result.sample_count == 3853
    assert result.source_raw_channel_count == 607
    assert result.report_channel_count == 326
    assert result.vsm_count == 187
    assert result.avl_count == 111
    assert result.math_count == 28
    assert result.statistic_count == 36
    assert result.kpi_count == 9
    assert result.plot_count == 20
    assert workbook["RoboSprayer Hybrid"].sheet_state == "visible"
    assert workbook["Rename From VSM to Astauto"].sheet_state == "visible"
    assert workbook["Metadata"].sheet_state == "hidden"
    assert len(workbook["RoboSprayer Hybrid"]._images) == 20


def test_profile_excel_report_exports_latest_electric_road_height(tmp_path: Path) -> None:
    result = generate_profile_excel_report(
        _latest_electric_csv(),
        ELECTRIC_PROFILE,
        tmp_path / "latest_electric",
        ImportOptions(),
    )
    workbook = load_workbook(result.report_path, data_only=True)
    sheet = workbook["RoboSprayer Electric"]
    by_semantic = {channel.channel_id: index + 1 for index, channel in enumerate(result.report_channels)}

    assert result.report_channel_count == 318
    assert result.plot_count == 14
    assert get_column_letter(result.report_channel_count) == "LF"
    assert by_semantic["track_height"] == by_semantic["track_gradient"] + 1
    assert sheet.cell(3, by_semantic["track_height"]).value == "Road Height"
    assert sheet.cell(4, by_semantic["track_height"]).value == "m"
    assert sheet.cell(DATA_START_ROW, by_semantic["track_height"]).value == pytest.approx(0.0)
    assert sheet.cell(DATA_START_ROW + 100, by_semantic["track_height"]).value > 0.0
    assert "road_profile" in [plot.plot_id for plot in result.plotting_result.rendered_plots]
    road_profile = result.plotting_result.series_summaries["road_profile"]
    assert [series.semantic_name for series in road_profile] == ["track_gradient", "track_height"]
    assert road_profile[1].unit == "m"


def test_profile_excel_report_exports_latest_hybrid_road_height(tmp_path: Path) -> None:
    result = generate_profile_excel_report(
        _latest_hybrid_csv(),
        HYBRID_PROFILE,
        tmp_path / "latest_hybrid",
        ImportOptions(),
    )
    workbook = load_workbook(result.report_path, data_only=True)
    sheet = workbook["Caiman SP Hybrid"]
    by_semantic = {channel.channel_id: index + 1 for index, channel in enumerate(result.report_channels)}

    assert result.report_channel_count == 326
    assert result.plot_count == 20
    assert result.report_path.name == "Caiman_SP_Hybrid_Engineering_Report.xlsx"
    assert result.report_metadata.report_title == "Caiman SP Hybrid"
    assert workbook.sheetnames[0] == "Caiman SP Hybrid"
    assert "track_height" not in by_semantic
    assert "Road Height" not in [sheet.cell(3, col).value for col in range(1, result.report_channel_count + 1)]
    assert "road_profile" in [plot.plot_id for plot in result.plotting_result.rendered_plots]
    road_profile = result.plotting_result.series_summaries["road_profile"]
    assert [series.semantic_name for series in road_profile] == ["track_gradient"]
    assert "wheel_loads" in [plot.plot_id for plot in result.plotting_result.rendered_plots]
    assert [series.semantic_name for series in result.plotting_result.series_summaries["wheel_loads"]] == [
        "wheel_load_dynamic_fl",
        "wheel_load_dynamic_fr",
        "wheel_load_dynamic_rl",
        "wheel_load_dynamic_rr",
    ]
    assert result.plotting_result.channels_by_semantic_name["driveshaft_torque_rl"].kind == "vsm"
    assert result.plotting_result.channels_by_semantic_name["driveshaft_torque_rr"].kind == "vsm"
    assert result.plotting_result.values_by_semantic_name["driveshaft_torque_rl"].max() > 14000.0
    assert result.plotting_result.values_by_semantic_name["driveshaft_torque_rr"].max() > 14000.0
    assert result.plotting_result.values_by_semantic_name["wheel_power_total"].max() > 100.0
    mapping = workbook["Rename From VSM to Astauto"]
    torque_row = by_semantic["driveshaft_torque_rl"] + 2
    assert mapping.cell(torque_row, 3).value == "VSM"
    assert (
        mapping.cell(torque_row, 8).value
        == "Raw VSM channel used when available; fallback = 0 when raw channel is unavailable."
    )
    metadata = {workbook["Metadata"].cell(row, 1).value: workbook["Metadata"].cell(row, 2).value for row in range(2, 45)}
    assert metadata["Resolved raw profile channels"] == (
        metadata["Exported VSM raw channels"] + metadata["Exported AVL raw channels"]
    )


def test_profile_excel_report_uses_dynamic_geometry_for_short_profile(tmp_path: Path) -> None:
    data_path = tmp_path / "short.csv"
    profile_path = tmp_path / "short_profile.yaml"
    data_path.write_text("Time,Speed\ns,kph\n0,0\n1,3\n2,6\n", encoding="utf-8")
    profile_path.write_text(
        "version: 1\n"
        "profile:\n"
        "  profile_id: short_profile\n"
        "  name: Short Profile\n"
        "  powertrain: electric\n"
        "channels:\n"
        "  raw:\n"
        "    - semantic_name: time\n"
        "      source_name: Time\n"
        "      report_name: Time (s)\n"
        "      unit: s\n"
        "      channel_type: VSM\n"
        "    - semantic_name: speed\n"
        "      source_name: Speed\n"
        "      report_name: Speed\n"
        "      unit: kph\n"
        "      channel_type: VSM\n"
        "  math:\n"
        "    - semantic_name: time_minutes\n"
        "      source_name: Time\n"
        "      report_name: Time (min)\n"
        "      unit: min\n"
        "      dependencies: [time]\n"
        "      expression: time / 60\n"
        "statistics:\n"
        "  - statistic_id: speed_max\n"
        "    target: speed\n"
        "    operation: max\n"
        "    display_name: Max Speed\n"
        "    unit: kph\n"
        "    placement_group: summary\n"
        "kpis:\n"
        "  - kpi_id: speed_summary\n"
        "    expression: speed_max\n"
        "    dependencies: [speed_max]\n"
        "    display_name: Speed Summary\n"
        "    unit: kph\n"
        "plots: []\n",
        encoding="utf-8",
    )

    result = generate_profile_excel_report(data_path, profile_path, tmp_path / "short_report", ImportOptions())
    workbook = load_workbook(result.report_path, data_only=True)
    sheet = workbook["Short Electric"]

    assert result.sample_count == 3
    assert result.report_channel_count == 3
    assert sheet.cell(5, 1).value == pytest.approx(0.0)
    assert sheet.cell(7, 1).value == pytest.approx(2.0)
    assert sheet.cell(8, 1).value == "MAX"
    assert sheet.cell(8, 2).value == pytest.approx(6.0)
    assert sheet.cell(3, 5).value == "Max Speed [kph]"
    assert sheet.cell(4, 5).value == pytest.approx(6.0)
    assert result.report_path.name == "Short_Electric_Engineering_Report.xlsx"
