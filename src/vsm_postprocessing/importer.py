from __future__ import annotations

import csv
import json
import math
import re
from dataclasses import dataclass
from itertools import zip_longest
from pathlib import Path
from statistics import median
from typing import Sequence

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .errors import DataValidationError, FileImportError
from .models import ChannelInfo, DataInspectionResult, DataQualityReport, ImportedDataset
from .utils import make_channel_id, normalized_name, sha256_file


@dataclass(frozen=True)
class ImportOptions:
    """Optional explicit overrides for safe import and inspection."""

    sheet_name: str | None = None
    header_row: int | None = None
    unit_row: int | None = None
    data_start_row: int | None = None
    data_end_row: int | None = None
    last_channel_column: int | None = None
    time_channel: str | None = None
    strict: bool = True


@dataclass
class _NumericScan:
    sample_count: int
    data_end_row: int
    time_values: np.ndarray
    missing_count: int
    invalid_count: int
    non_finite_count: int
    invalid_examples: list[str]
    formula_metadata: dict[int, tuple[str | None, tuple[str, ...]]]


def inspect_data_file(path: str | Path, options: ImportOptions | None = None) -> DataInspectionResult:
    source_path = Path(path).expanduser().resolve()
    options = options or ImportOptions()

    if not source_path.exists():
        raise FileImportError(f"Input file does not exist: {source_path}")
    if not source_path.is_file():
        raise FileImportError(f"Input path is not a file: {source_path}")

    suffix = source_path.suffix.lower()
    if suffix == ".xlsx":
        result = _inspect_xlsx(source_path, options)
    elif suffix == ".csv":
        result = _inspect_csv(source_path, options)
    else:
        raise FileImportError(f"Unsupported input type '{suffix}'. Supported types are .xlsx and .csv")

    if options.strict and not result.quality.is_valid:
        raise DataValidationError("; ".join(result.quality.errors))
    return result


def load_data_file(path: str | Path, options: ImportOptions | None = None) -> ImportedDataset:
    """Load a fully validated numeric XLSX/CSV dataset into a float64 matrix."""

    source_path = Path(path).expanduser().resolve()
    options = options or ImportOptions()
    inspection = inspect_data_file(source_path, options)
    quality = inspection.quality

    if source_path.suffix.lower() == ".xlsx":
        values = _load_xlsx_numeric_matrix(source_path, quality)
    elif source_path.suffix.lower() == ".csv":
        values = _load_csv_numeric_matrix(source_path, quality)
    else:  # inspect_data_file has already validated the suffix
        raise FileImportError(f"Unsupported input type '{source_path.suffix.lower()}'")

    return ImportedDataset(
        source_path=source_path,
        channels=inspection.channels,
        quality=quality,
        values=values,
    )


def _load_xlsx_numeric_matrix(path: Path, quality: DataQualityReport) -> np.ndarray:
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover
        raise FileImportError(f"Could not open Excel workbook '{path}' for data loading: {exc}") from exc

    try:
        if quality.sheet_name is None or quality.sheet_name not in workbook.sheetnames:
            raise FileImportError(f"Validated sheet '{quality.sheet_name}' is not available in '{path}'")
        sheet = workbook[quality.sheet_name]
        values = np.empty((quality.sample_count, quality.channel_count), dtype=np.float64)
        rows = sheet.iter_rows(
            min_row=quality.data_start_row,
            max_row=quality.data_end_row,
            min_col=1,
            max_col=quality.channel_count,
            values_only=True,
        )
        row_count = 0
        for row_index, row in enumerate(rows):
            for column_index, raw_value in enumerate(row):
                parsed, status = _coerce_numeric(raw_value)
                if status != "ok":
                    coordinate = f"{get_column_letter(column_index + 1)}{quality.data_start_row + row_index}"
                    raise DataValidationError(
                        f"Validated data changed or could not be loaded at {coordinate}: {raw_value!r} ({status})"
                    )
                values[row_index, column_index] = parsed
            row_count += 1
        if row_count != quality.sample_count:
            raise DataValidationError(
                f"Expected {quality.sample_count} data rows but loaded {row_count} from '{path}'"
            )
        return values
    finally:
        workbook.close()


def _load_csv_numeric_matrix(path: Path, quality: DataQualityReport) -> np.ndarray:
    rows = _read_csv_rows(path)
    data_rows = rows[quality.data_start_row - 1 : quality.data_end_row]
    if len(data_rows) != quality.sample_count:
        raise DataValidationError(
            f"Expected {quality.sample_count} CSV data rows but loaded {len(data_rows)} from '{path}'"
        )

    values = np.empty((quality.sample_count, quality.channel_count), dtype=np.float64)
    for row_index, source_row in enumerate(data_rows):
        padded = list(source_row[: quality.channel_count]) + [None] * max(
            0, quality.channel_count - len(source_row)
        )
        for column_index, raw_value in enumerate(padded):
            parsed, status = _coerce_numeric(raw_value)
            if status != "ok":
                raise DataValidationError(
                    "Validated CSV data changed or could not be loaded at "
                    f"row {quality.data_start_row + row_index}, column {column_index + 1}: "
                    f"{raw_value!r} ({status})"
                )
            values[row_index, column_index] = parsed
    return values


def _inspect_xlsx(path: Path, options: ImportOptions) -> DataInspectionResult:
    try:
        formula_book = load_workbook(path, data_only=False, read_only=True)
        value_book = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover
        raise FileImportError(f"Could not open Excel workbook '{path}': {exc}") from exc

    try:
        sheet_name = options.sheet_name or formula_book.sheetnames[0]
        if sheet_name not in formula_book.sheetnames:
            available = ", ".join(formula_book.sheetnames)
            raise FileImportError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")

        formula_sheet = formula_book[sheet_name]
        value_sheet = value_book[sheet_name]
        preview = list(
            formula_sheet.iter_rows(
                min_row=1,
                max_row=min(formula_sheet.max_row, 30),
                values_only=True,
            )
        )
        header_row = options.header_row or _detect_header_row_from_rows(preview)
        headers_full = list(_get_preview_or_streamed_row(formula_sheet, preview, header_row))
        if options.last_channel_column is not None:
            if options.last_channel_column < 1:
                raise FileImportError("last_channel_column must be at least 1")
            last_channel_column = options.last_channel_column
            if last_channel_column > len(headers_full):
                headers_full.extend([None] * (last_channel_column - len(headers_full)))
        else:
            last_channel_column = _last_nonempty_index(headers_full)
        if last_channel_column == 0:
            raise FileImportError(f"No channel names found in detected header row {header_row}")
        headers = headers_full[:last_channel_column]
        _validate_headers(headers, header_row)

        unit_row = options.unit_row
        if unit_row is None:
            candidate = header_row + 1
            candidate_values = list(_get_preview_or_streamed_row(formula_sheet, preview, candidate))[:last_channel_column]
            if _looks_like_unit_row(candidate_values, headers):
                unit_row = candidate

        units = (
            [
                _clean_optional_text(value)
                for value in list(_get_preview_or_streamed_row(formula_sheet, preview, unit_row))[:last_channel_column]
            ]
            if unit_row is not None
            else [None] * last_channel_column
        )
        data_start_row = options.data_start_row or ((unit_row or header_row) + 1)
        if options.data_end_row is not None and options.data_end_row < data_start_row:
            raise FileImportError(
                f"data_end_row ({options.data_end_row}) must be greater than or equal to data_start_row ({data_start_row})"
            )
        channel_ids = [make_channel_id(str(name), index) for index, name in enumerate(headers, start=1)]
        time_column_index = _detect_time_column(headers, units, options.time_channel)

        scan = _scan_xlsx_data(
            value_sheet=value_sheet,
            formula_sheet=formula_sheet,
            data_start_row=data_start_row,
            last_column=last_channel_column,
            time_column_index=time_column_index,
            channel_ids=channel_ids,
            requested_data_end_row=options.data_end_row,
        )
        if scan.sample_count == 0:
            raise FileImportError(f"No data rows found from row {data_start_row} onward")

        channels = _build_channels(
            path=path,
            sheet_name=sheet_name,
            headers=headers,
            units=units,
            channel_ids=channel_ids,
            formula_metadata=scan.formula_metadata,
        )
        quality = _build_quality_report(
            path=path,
            file_type="xlsx",
            sheet_name=sheet_name,
            header_row=header_row,
            unit_row=unit_row,
            data_start_row=data_start_row,
            data_end_row=scan.data_end_row,
            sample_count=scan.sample_count,
            time_values=scan.time_values,
            channels=channels,
            time_column_index=time_column_index,
            missing_count=scan.missing_count,
            invalid_count=scan.invalid_count,
            non_finite_count=scan.non_finite_count,
            invalid_examples=scan.invalid_examples,
        )
        return DataInspectionResult(source_path=path, channels=channels, quality=quality)
    finally:
        formula_book.close()
        value_book.close()


def _scan_xlsx_data(
    value_sheet,
    formula_sheet,
    data_start_row: int,
    last_column: int,
    time_column_index: int,
    channel_ids: Sequence[str],
    requested_data_end_row: int | None = None,
) -> _NumericScan:
    max_data_row = requested_data_end_row or value_sheet.max_row
    if max_data_row > value_sheet.max_row:
        raise FileImportError(
            f"Requested data_end_row {max_data_row} exceeds worksheet maximum row {value_sheet.max_row}"
        )
    value_rows = value_sheet.iter_rows(
        min_row=data_start_row,
        max_row=max_data_row,
        min_col=1,
        max_col=last_column,
        values_only=True,
    )
    formula_rows = formula_sheet.iter_rows(
        min_row=data_start_row,
        max_row=max_data_row,
        min_col=1,
        max_col=last_column,
        values_only=True,
    )

    first_formulas: list[str | None] = [None] * last_column
    dependencies: list[list[str]] = [[] for _ in range(last_column)]
    reference_pattern = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?\d+")
    time_values: list[float] = []
    missing_count = 0
    invalid_count = 0
    non_finite_count = 0
    invalid_examples: list[str] = []
    sample_count = 0

    for row_offset, (value_row, formula_row) in enumerate(zip_longest(value_rows, formula_rows, fillvalue=())):
        row_index = data_start_row + row_offset
        value_row = tuple(value_row) + (None,) * max(0, last_column - len(value_row))
        formula_row = tuple(formula_row) + (None,) * max(0, last_column - len(formula_row))

        raw_time = value_row[time_column_index - 1]
        if raw_time is None or (isinstance(raw_time, str) and not raw_time.strip()):
            break

        sample_count += 1
        for column_offset in range(last_column):
            cached_value = value_row[column_offset]
            formula_value = formula_row[column_offset]

            if first_formulas[column_offset] is None and isinstance(formula_value, str) and formula_value.startswith("="):
                first_formulas[column_offset] = formula_value
                for letters in reference_pattern.findall(formula_value.upper()):
                    referenced_index = _column_letters_to_index(letters)
                    if 1 <= referenced_index <= len(channel_ids):
                        referenced_id = channel_ids[referenced_index - 1]
                        if referenced_id not in dependencies[column_offset]:
                            dependencies[column_offset].append(referenced_id)

            parsed, status = _coerce_numeric(cached_value)
            if status == "missing" and isinstance(formula_value, str) and formula_value.startswith("="):
                status = "invalid"
            if status == "missing":
                missing_count += 1
            elif status == "invalid":
                invalid_count += 1
                if len(invalid_examples) < 10:
                    coordinate = f"{get_column_letter(column_offset + 1)}{row_index}"
                    invalid_examples.append(f"{coordinate}={cached_value!r}")
            elif status == "non_finite":
                non_finite_count += 1

            if column_offset == time_column_index - 1:
                time_values.append(parsed)

    formula_metadata = {
        index + 1: (first_formulas[index], tuple(dependencies[index]))
        for index in range(last_column)
    }
    return _NumericScan(
        sample_count=sample_count,
        data_end_row=data_start_row + sample_count - 1,
        time_values=np.asarray(time_values, dtype=float),
        missing_count=missing_count,
        invalid_count=invalid_count,
        non_finite_count=non_finite_count,
        invalid_examples=invalid_examples,
        formula_metadata=formula_metadata,
    )


def _inspect_csv(path: Path, options: ImportOptions) -> DataInspectionResult:
    rows = _read_csv_rows(path)
    if not rows:
        raise FileImportError(f"CSV file is empty: {path}")

    header_row = options.header_row or _detect_header_row_from_rows(rows)
    headers_raw = rows[header_row - 1]
    if options.last_channel_column is not None:
        if options.last_channel_column < 1:
            raise FileImportError("last_channel_column must be at least 1")
        last_channel_column = options.last_channel_column
    else:
        last_channel_column = _last_nonempty_index(headers_raw)
    headers = list(headers_raw[:last_channel_column]) + [None] * max(0, last_channel_column - len(headers_raw))
    _validate_headers(headers, header_row)

    unit_row = options.unit_row
    if unit_row is None and header_row < len(rows):
        candidate = header_row + 1
        if _looks_like_unit_row(rows[candidate - 1][:last_channel_column], headers):
            unit_row = candidate

    units = (
        [_clean_optional_text(value) for value in rows[unit_row - 1][:last_channel_column]]
        if unit_row is not None
        else [None] * last_channel_column
    )
    units.extend([None] * max(0, last_channel_column - len(units)))
    data_start_row = options.data_start_row or ((unit_row or header_row) + 1)
    if options.data_end_row is not None and options.data_end_row < data_start_row:
        raise FileImportError(
            f"data_end_row ({options.data_end_row}) must be greater than or equal to data_start_row ({data_start_row})"
        )
    time_column_index = _detect_time_column(headers, units, options.time_channel)

    data_rows = rows[data_start_row - 1 : options.data_end_row]
    scan = _scan_csv_data(data_rows, last_channel_column, time_column_index, data_start_row)
    if scan.sample_count == 0:
        raise FileImportError(f"No CSV data rows found from row {data_start_row} onward")

    channel_ids = [make_channel_id(str(name), index) for index, name in enumerate(headers, start=1)]
    channels = [
        ChannelInfo(
            channel_id=channel_ids[index - 1],
            source_name=str(name).strip(),
            display_name=str(name).strip(),
            unit=units[index - 1],
            source_column_index=index,
            source_column_label=str(index),
            kind="raw",
            dtype="float64",
            provenance=f"{path.name}:CSV column {index}",
        )
        for index, name in enumerate(headers, start=1)
    ]
    quality = _build_quality_report(
        path=path,
        file_type="csv",
        sheet_name=None,
        header_row=header_row,
        unit_row=unit_row,
        data_start_row=data_start_row,
        data_end_row=scan.data_end_row,
        sample_count=scan.sample_count,
        time_values=scan.time_values,
        channels=channels,
        time_column_index=time_column_index,
        missing_count=scan.missing_count,
        invalid_count=scan.invalid_count,
        non_finite_count=scan.non_finite_count,
        invalid_examples=scan.invalid_examples,
    )
    return DataInspectionResult(source_path=path, channels=channels, quality=quality)


def _scan_csv_data(
    rows: Sequence[Sequence[object]],
    last_column: int,
    time_column_index: int,
    data_start_row: int,
) -> _NumericScan:
    time_values: list[float] = []
    missing_count = 0
    invalid_count = 0
    non_finite_count = 0
    invalid_examples: list[str] = []
    sample_count = 0

    for row_offset, source_row in enumerate(rows):
        padded = list(source_row[:last_column]) + [None] * max(0, last_column - len(source_row))
        raw_time = padded[time_column_index - 1]
        if raw_time is None or (isinstance(raw_time, str) and not raw_time.strip()):
            break
        sample_count += 1
        for column_index, raw_value in enumerate(padded, start=1):
            parsed, status = _coerce_numeric(raw_value)
            if status == "missing":
                missing_count += 1
            elif status == "invalid":
                invalid_count += 1
                if len(invalid_examples) < 10:
                    invalid_examples.append(
                        f"row {data_start_row + row_offset}, column {column_index}={raw_value!r}"
                    )
            elif status == "non_finite":
                non_finite_count += 1
            if column_index == time_column_index:
                time_values.append(parsed)

    return _NumericScan(
        sample_count=sample_count,
        data_end_row=data_start_row + sample_count - 1,
        time_values=np.asarray(time_values, dtype=float),
        missing_count=missing_count,
        invalid_count=invalid_count,
        non_finite_count=non_finite_count,
        invalid_examples=invalid_examples,
        formula_metadata={},
    )


def _detect_header_row_from_rows(rows: Sequence[Sequence[object]]) -> int:
    candidates: list[tuple[float, int]] = []
    for row_index, values in enumerate(rows[:30], start=1):
        score = _header_score(values)
        if score > 0:
            candidates.append((score, row_index))
    if not candidates:
        raise FileImportError("Could not auto-detect a header row in the first 30 rows")
    return max(candidates)[1]


def _header_score(values: Sequence[object]) -> float:
    nonempty = [str(value).strip() for value in values if _clean_optional_text(value) is not None]
    if len(nonempty) < 2:
        return -1.0
    text_count = sum(not _looks_numeric_text(value) and not value.startswith("=") for value in nonempty)
    text_ratio = text_count / len(nonempty)
    if text_ratio < 0.70:
        return -1.0
    unique_ratio = len(set(nonempty)) / len(nonempty)
    average_length = sum(len(value) for value in nonempty) / len(nonempty)
    return len(nonempty) + 20.0 * unique_ratio + min(average_length, 40.0) / 10.0


def _looks_like_unit_row(values: Sequence[object], headers: Sequence[object]) -> bool:
    cleaned = [str(value).strip() for value in values if _clean_optional_text(value) is not None]
    if len(cleaned) < max(1, int(0.5 * len(headers))):
        return False
    numeric_like = sum(_looks_numeric_text(value) or value.startswith("=") for value in cleaned)
    unique_ratio = len(set(cleaned)) / len(cleaned)
    average_length = sum(len(value) for value in cleaned) / len(cleaned)
    return numeric_like == 0 and average_length <= 12 and (unique_ratio < 0.75 or average_length <= 5)


def _detect_time_column(headers: Sequence[object], units: Sequence[str | None], requested: str | None) -> int:
    names = [str(value).strip() for value in headers]
    normalized = [normalized_name(name) for name in names]

    if requested:
        requested_normalized = normalized_name(requested)
        matches = [index for index, name in enumerate(normalized, start=1) if name == requested_normalized]
        if not matches:
            raise FileImportError(f"Requested time channel '{requested}' was not found in the header")
        if len(matches) > 1:
            raise FileImportError(f"Requested time channel '{requested}' is ambiguous because it appears more than once")
        return matches[0]

    for preferred in ("tracktime", "simulationtime", "simtime", "timestamp", "elapsedtime", "time"):
        for index, name in enumerate(normalized, start=1):
            if name == preferred:
                return index

    second_units = {"s", "sec", "secs", "second", "seconds"}
    unit_matches = [
        index
        for index, unit in enumerate(units, start=1)
        if unit is not None and normalized_name(unit) in second_units
    ]
    if len(unit_matches) == 1:
        return unit_matches[0]
    raise FileImportError("Could not identify a unique time channel. Use --time-channel to specify it explicitly")


def _coerce_numeric(value: object) -> tuple[float, str]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return math.nan, "missing"
    if isinstance(value, bool):
        return float(value), "ok"
    if isinstance(value, (int, float, np.number)):
        number = float(value)
    elif isinstance(value, str):
        text = value.strip().replace(" ", "")
        try:
            number = float(text)
        except ValueError:
            if text.count(",") == 1 and "." not in text:
                try:
                    number = float(text.replace(",", "."))
                except ValueError:
                    return math.nan, "invalid"
            else:
                return math.nan, "invalid"
    else:
        return math.nan, "invalid"

    if not math.isfinite(number):
        return number, "non_finite"
    return number, "ok"


def _build_channels(
    path: Path,
    sheet_name: str,
    headers: Sequence[object],
    units: Sequence[str | None],
    channel_ids: Sequence[str],
    formula_metadata: dict[int, tuple[str | None, tuple[str, ...]]],
) -> list[ChannelInfo]:
    channels: list[ChannelInfo] = []
    for index, raw_name in enumerate(headers, start=1):
        source_name = str(raw_name).strip()
        formula_example, dependencies = formula_metadata[index]
        channels.append(
            ChannelInfo(
                channel_id=channel_ids[index - 1],
                source_name=source_name,
                display_name=source_name,
                unit=units[index - 1],
                source_column_index=index,
                source_column_label=get_column_letter(index),
                kind="math" if formula_example is not None else "raw",
                dtype="float64",
                provenance=f"{path.name}:{sheet_name}!{get_column_letter(index)}",
                dependencies=dependencies,
                formula_example=formula_example,
            )
        )
    return channels


def _build_quality_report(
    path: Path,
    file_type: str,
    sheet_name: str | None,
    header_row: int,
    unit_row: int | None,
    data_start_row: int,
    data_end_row: int,
    sample_count: int,
    time_values: np.ndarray,
    channels: Sequence[ChannelInfo],
    time_column_index: int,
    missing_count: int,
    invalid_count: int,
    non_finite_count: int,
    invalid_examples: Sequence[str],
) -> DataQualityReport:
    errors: list[str] = []
    warnings: list[str] = []
    finite_time = np.isfinite(time_values)
    if not finite_time.all():
        errors.append(f"Time channel contains {int((~finite_time).sum())} missing, invalid or non-finite values")

    valid_time = time_values[finite_time]
    if valid_time.size >= 2:
        differences = np.diff(valid_time)
        duplicate_count = int(np.sum(differences == 0.0))
        strictly_increasing = bool(np.all(differences > 0.0))
        nominal_step = float(median(float(value) for value in differences))
        time_start = float(valid_time[0])
        time_end = float(valid_time[-1])
        tolerance = max(abs(nominal_step) * 1e-9, 1e-12)
        irregular_count = int(np.sum(np.abs(differences - nominal_step) > tolerance))
        if not strictly_increasing:
            errors.append("Time channel is not strictly increasing")
        if duplicate_count:
            errors.append(f"Time channel contains {duplicate_count} duplicate timestamps")
        if irregular_count:
            warnings.append(f"Time step is not uniform in {irregular_count} intervals")
    elif valid_time.size == 1:
        duplicate_count = 0
        strictly_increasing = True
        nominal_step = None
        time_start = time_end = float(valid_time[0])
        warnings.append("Only one valid time sample is available; time-step checks were skipped")
    else:
        duplicate_count = 0
        strictly_increasing = False
        nominal_step = None
        time_start = time_end = None

    if missing_count:
        errors.append(f"Data region contains {missing_count} missing cells")
    if invalid_count:
        errors.append(
            f"Data region contains {invalid_count} nonnumeric cells. Examples: {', '.join(invalid_examples)}"
        )
    if non_finite_count:
        errors.append(f"Data region contains {non_finite_count} non-finite numeric cells")

    missing_units = sum(channel.unit is None for channel in channels)
    if missing_units:
        warnings.append(f"Units are missing for {missing_units} channels")

    time_channel = channels[time_column_index - 1]
    raw_count = sum(channel.kind == "raw" for channel in channels)
    math_count = sum(channel.kind == "math" for channel in channels)
    return DataQualityReport(
        source_file=str(path),
        source_sha256=sha256_file(path),
        file_type=file_type,
        sheet_name=sheet_name,
        header_row=header_row,
        unit_row=unit_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        sample_count=sample_count,
        channel_count=len(channels),
        raw_channel_count=raw_count,
        math_channel_count=math_count,
        time_channel_id=time_channel.channel_id,
        time_channel_name=time_channel.source_name,
        time_unit=time_channel.unit,
        time_start=time_start,
        time_end=time_end,
        nominal_time_step=nominal_step,
        time_is_strictly_increasing=strictly_increasing,
        duplicate_timestamp_count=duplicate_count,
        missing_cell_count=missing_count,
        invalid_numeric_cell_count=invalid_count,
        non_finite_cell_count=non_finite_count,
        warnings=warnings,
        errors=errors,
    )


def export_inspection(result: DataInspectionResult, output_dir: str | Path) -> dict[str, Path]:
    output_path = Path(output_dir).expanduser().resolve()
    output_path.mkdir(parents=True, exist_ok=True)

    catalogue_path = output_path / "channel_catalogue.csv"
    quality_path = output_path / "data_quality_report.json"
    inspection_path = output_path / "inspection_result.json"
    summary_path = output_path / "inspection_summary.txt"

    with catalogue_path.open("w", newline="", encoding="utf-8") as handle:
        fieldnames = [
            "channel_id",
            "source_name",
            "display_name",
            "unit",
            "source_column_index",
            "source_column_label",
            "kind",
            "dtype",
            "provenance",
            "dependencies",
            "formula_example",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for channel in result.channels:
            row = channel.to_dict()
            row["dependencies"] = ";".join(row["dependencies"])
            writer.writerow(row)

    quality_path.write_text(json.dumps(result.quality.to_dict(), indent=2), encoding="utf-8")
    inspection_path.write_text(json.dumps(result.to_dict(), indent=2), encoding="utf-8")
    summary_path.write_text(_format_summary(result), encoding="utf-8")
    return {
        "channel_catalogue": catalogue_path,
        "data_quality_report": quality_path,
        "inspection_result": inspection_path,
        "inspection_summary": summary_path,
    }


def _format_summary(result: DataInspectionResult) -> str:
    report = result.quality
    lines = [
        "VSM DATA INSPECTION",
        "===================",
        f"Status: {'PASS' if report.is_valid else 'FAIL'}",
        f"Source: {report.source_file}",
        f"SHA-256: {report.source_sha256}",
        f"Sheet: {report.sheet_name or '-'}",
        f"Header row: {report.header_row}",
        f"Unit row: {report.unit_row or '-'}",
        f"Data rows: {report.data_start_row}-{report.data_end_row}",
        f"Samples: {report.sample_count}",
        f"Channels: {report.channel_count} ({report.raw_channel_count} raw, {report.math_channel_count} math)",
        f"Time channel: {report.time_channel_name} [{report.time_unit or '-'}]",
        f"Time range: {report.time_start} to {report.time_end}",
        f"Nominal time step: {report.nominal_time_step}",
        f"Strictly increasing time: {report.time_is_strictly_increasing}",
        f"Duplicate timestamps: {report.duplicate_timestamp_count}",
        f"Missing cells: {report.missing_cell_count}",
        f"Invalid numeric cells: {report.invalid_numeric_cell_count}",
        f"Non-finite cells: {report.non_finite_cell_count}",
    ]
    if report.warnings:
        lines.extend(["", "Warnings:", *[f"- {warning}" for warning in report.warnings]])
    if report.errors:
        lines.extend(["", "Errors:", *[f"- {error}" for error in report.errors]])
    return "\n".join(lines) + "\n"


def _read_csv_rows(path: Path) -> list[list[str]]:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            sample = path.read_text(encoding=encoding)[:8192]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            with path.open("r", newline="", encoding=encoding) as handle:
                return [list(row) for row in csv.reader(handle, dialect)]
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            raise FileImportError(f"Could not read CSV file '{path}': {exc}") from exc
    raise FileImportError(f"Could not decode CSV file '{path}'")


def _get_preview_or_streamed_row(sheet, preview: Sequence[Sequence[object]], row_index: int) -> Sequence[object]:
    if 1 <= row_index <= len(preview):
        return preview[row_index - 1]
    rows = sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True)
    return next(rows, ())


def _last_nonempty_index(values: Sequence[object]) -> int:
    for index in range(len(values), 0, -1):
        if _clean_optional_text(values[index - 1]) is not None:
            return index
    return 0


def _validate_headers(headers: Sequence[object], header_row: int) -> None:
    missing = [index for index, value in enumerate(headers, start=1) if _clean_optional_text(value) is None]
    if missing:
        raise FileImportError(f"Header row {header_row} contains blank channel names at columns {missing}")


def _clean_optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _looks_numeric_text(value: str) -> bool:
    try:
        float(value.replace(",", "."))
        return True
    except ValueError:
        return False


def _column_letters_to_index(letters: str) -> int:
    result = 0
    for character in letters:
        result = result * 26 + (ord(character) - ord("A") + 1)
    return result
