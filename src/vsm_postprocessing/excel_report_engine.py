from __future__ import annotations

import json
import math
import shutil
import csv
from datetime import datetime, timezone
from dataclasses import replace
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping

import numpy as np
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .errors import ConfigurationError, ExcelReportError
from .importer import ImportOptions, load_data_file
from .models import ChannelInfo
from .plotting_engine import PlottingResult, load_plotting_config, render_plots
from .profile_math import ProfileMathResult, calculate_profile_math_channels
from .profile_plotting import ProfilePlottingResult, render_profile_plots
from .profile_statistics import ProfileStatisticsResult, calculate_profile_statistics
from .report_profile import ProfileResolutionResult, ReportingProfile, load_reporting_profile, resolve_profile
from .statistics_engine import StatisticResult, StatisticsResult, calculate_statistics
from .utils import client_display_filename, normalized_name, sha256_file
from .version import __version__

_ALLOWED_BOTTOM_OPERATIONS = ("max", "min", "last", "sum", "rms", "time_weighted_rms")


@dataclass(frozen=True)
class ExcelReportConfig:
    version: int
    title: str
    subtitle: str | None
    report_sheet: str
    metadata_sheet: str
    channel_ids: tuple[str, ...]
    channel_metadata: Mapping[str, Mapping[str, str]]
    top_rms_statistic_ids: tuple[str, ...]
    kpi_statistic_ids: tuple[str, ...]
    bottom_operations: tuple[str, ...]
    bottom_summary_statistic_ids: tuple[str, ...]
    layout_profile: str
    plot_placement: str
    blank_separator_columns: int
    channel_width: int
    header_row_height: int
    unit_row_height: int
    plot_ids: tuple[str, ...]
    plot_columns: int
    plot_width_px: int
    plot_height_px: int
    native_chart_ids: tuple[str, ...]
    chart_layout: Mapping[str, Mapping[str, Any]]
    rms_merges: Mapping[str, str]
    output_filename: str
    keep_plot_assets: bool


@dataclass
class ExcelReportResult:
    report_path: Path
    manifest_path: Path
    summary_path: Path
    plot_assets_dir: Path
    config_path: Path
    config: ExcelReportConfig
    statistics_result: StatisticsResult
    plotting_result: PlottingResult
    report_channels: list[ChannelInfo]

    @property
    def sample_count(self) -> int:
        return self.statistics_result.sample_count

    @property
    def channel_count(self) -> int:
        return len(self.report_channels)

    @property
    def statistic_count(self) -> int:
        return self.statistics_result.statistic_count

    @property
    def plot_count(self) -> int:
        return self.plotting_result.plot_count

    @property
    def configured_plot_count(self) -> int:
        return self.plotting_result.plot_count

    @property
    def native_excel_chart_count(self) -> int:
        return len(self.config.native_chart_ids)

    @property
    def embedded_plot_image_count(self) -> int:
        return len(self.config.plot_ids)


@dataclass
class ProfileExcelReportResult:
    report_path: Path
    manifest_path: Path
    summary_path: Path
    plot_assets_dir: Path
    dataset: Any
    profile: ReportingProfile
    resolution: ProfileResolutionResult
    math_result: ProfileMathResult
    statistics_result: ProfileStatisticsResult
    plotting_result: ProfilePlottingResult
    report_channels: list[ChannelInfo]
    template_comparison: list[dict[str, str]]

    @property
    def sample_count(self) -> int:
        return self.dataset.quality.sample_count

    @property
    def source_raw_channel_count(self) -> int:
        return self.dataset.quality.raw_channel_count

    @property
    def report_channel_count(self) -> int:
        return len(self.report_channels)

    @property
    def vsm_count(self) -> int:
        raw_by_name = self.profile.raw_by_semantic_name()
        return sum(1 for channel in self.report_channels if raw_by_name.get(channel.channel_id) and raw_by_name[channel.channel_id].channel_type == "VSM")

    @property
    def avl_count(self) -> int:
        raw_by_name = self.profile.raw_by_semantic_name()
        return sum(1 for channel in self.report_channels if raw_by_name.get(channel.channel_id) and raw_by_name[channel.channel_id].channel_type == "AVL")

    @property
    def math_count(self) -> int:
        return len(self.math_result.calculated_channels)

    @property
    def statistic_count(self) -> int:
        return self.statistics_result.calculated_statistic_count

    @property
    def kpi_count(self) -> int:
        return self.statistics_result.calculated_kpi_count

    @property
    def plot_count(self) -> int:
        return self.plotting_result.rendered_plot_count


def load_excel_report_config(path: str | Path) -> ExcelReportConfig:
    config_path = Path(path).expanduser().resolve()
    if not config_path.exists():
        raise ConfigurationError(f"Excel-report configuration does not exist: {config_path}")
    if not config_path.is_file():
        raise ConfigurationError(f"Excel-report configuration is not a file: {config_path}")

    try:
        raw = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    except UnicodeDecodeError as exc:
        raise ConfigurationError(f"Excel-report configuration must be UTF-8: {config_path}") from exc
    except yaml.YAMLError as exc:
        raise ConfigurationError(f"Invalid YAML in Excel-report configuration '{config_path}': {exc}") from exc

    if not isinstance(raw, dict):
        raise ConfigurationError("Excel-report configuration root must be a YAML mapping")
    _reject_unknown_keys(
        raw,
        {"version", "report", "channels", "channel_metadata", "statistics", "layout", "plots", "output"},
        "root",
    )

    version = raw.get("version")
    if version != 1:
        raise ConfigurationError("Excel-report configuration 'version' must be 1")

    report_raw = raw.get("report", {})
    if not isinstance(report_raw, dict):
        raise ConfigurationError("report must be a YAML mapping")
    _reject_unknown_keys(report_raw, {"title", "subtitle", "report_sheet", "metadata_sheet"}, "report")
    title = _nonempty_string(report_raw.get("title", "VSM Engineering Report"), "report.title")
    subtitle = _optional_string(report_raw.get("subtitle"), "report.subtitle")
    report_sheet = _sheet_name(report_raw.get("report_sheet", "Report"), "report.report_sheet")
    metadata_sheet = _sheet_name(report_raw.get("metadata_sheet", "Metadata"), "report.metadata_sheet")
    if report_sheet == metadata_sheet:
        raise ConfigurationError("report.report_sheet and report.metadata_sheet must be different")

    channels_raw = raw.get("channels")
    if not isinstance(channels_raw, list) or not channels_raw:
        raise ConfigurationError("channels must be a non-empty YAML list of channel_id values")
    channel_ids = tuple(_nonempty_string(value, "channels[]") for value in channels_raw)
    duplicates = _duplicates(channel_ids)
    if duplicates:
        raise ConfigurationError("channels contains duplicate channel IDs: " + ", ".join(duplicates))
    channel_metadata = _load_channel_metadata(raw.get("channel_metadata", {}), channel_ids)

    statistics_raw = raw.get("statistics", {})
    if not isinstance(statistics_raw, dict):
        raise ConfigurationError("statistics must be a YAML mapping")
    _reject_unknown_keys(
        statistics_raw,
        {"top_rms", "kpis", "bottom_operations", "bottom_summary"},
        "statistics",
    )
    top_rms = _string_list(statistics_raw.get("top_rms", []), "statistics.top_rms")
    kpis = _string_list(statistics_raw.get("kpis", []), "statistics.kpis")
    bottom_raw = statistics_raw.get("bottom_operations", list(_ALLOWED_BOTTOM_OPERATIONS))
    bottom_operations = _string_list(bottom_raw, "statistics.bottom_operations")
    invalid_ops = sorted(set(bottom_operations) - set(_ALLOWED_BOTTOM_OPERATIONS))
    if invalid_ops:
        raise ConfigurationError(
            "statistics.bottom_operations contains unsupported operations: " + ", ".join(invalid_ops)
        )
    if len(set(bottom_operations)) != len(bottom_operations):
        raise ConfigurationError("statistics.bottom_operations contains duplicates")
    bottom_summary = _string_list(statistics_raw.get("bottom_summary", []), "statistics.bottom_summary")

    layout_raw = raw.get("layout", {})
    if not isinstance(layout_raw, dict):
        raise ConfigurationError("layout must be a YAML mapping")
    _reject_unknown_keys(
        layout_raw,
        {
            "profile",
            "plot_placement",
            "blank_separator_columns",
            "channel_width",
            "header_row_height",
            "unit_row_height",
            "rms_merges",
        },
        "layout",
    )
    layout_profile = _choice(
        layout_raw.get("profile", "engineering"),
        "layout.profile",
        {"engineering", "sergio_reference"},
    )
    default_plot_placement = "kpi_panel" if layout_profile == "sergio_reference" else "below_data"
    plot_placement = _choice(
        layout_raw.get("plot_placement", default_plot_placement),
        "layout.plot_placement",
        {"below_data", "kpi_panel"},
    )
    blank_separator_columns = _positive_int(
        layout_raw.get("blank_separator_columns", 1),
        "layout.blank_separator_columns",
        minimum=1,
        maximum=4,
    )
    channel_width = _positive_int(layout_raw.get("channel_width", 12), "layout.channel_width", minimum=7, maximum=30)
    header_row_height = _positive_int(
        layout_raw.get("header_row_height", 106 if layout_profile == "sergio_reference" else 44),
        "layout.header_row_height",
        minimum=24,
        maximum=180,
    )
    unit_row_height = _positive_int(
        layout_raw.get("unit_row_height", 16 if layout_profile == "sergio_reference" else 26),
        "layout.unit_row_height",
        minimum=12,
        maximum=60,
    )
    rms_merges = _load_string_mapping(layout_raw.get("rms_merges", {}), "layout.rms_merges")

    plots_raw = raw.get("plots", {})
    if not isinstance(plots_raw, dict):
        raise ConfigurationError("plots must be a YAML mapping")
    _reject_unknown_keys(plots_raw, {"include", "native_charts", "chart_layout", "columns", "width_px", "height_px"}, "plots")
    plot_ids = _string_list(plots_raw.get("include", []), "plots.include")
    native_chart_ids = _string_list(plots_raw.get("native_charts", []), "plots.native_charts")
    chart_layout = _load_chart_layout(plots_raw.get("chart_layout", {}))
    plot_columns = _positive_int(plots_raw.get("columns", 2), "plots.columns", maximum=4)
    plot_width_px = _positive_int(plots_raw.get("width_px", 720), "plots.width_px", minimum=240, maximum=1600)
    plot_height_px = _positive_int(plots_raw.get("height_px", 360), "plots.height_px", minimum=160, maximum=1000)

    output_raw = raw.get("output", {})
    if not isinstance(output_raw, dict):
        raise ConfigurationError("output must be a YAML mapping")
    _reject_unknown_keys(output_raw, {"filename", "keep_plot_assets"}, "output")
    output_filename = _plain_xlsx_filename(output_raw.get("filename", "vsm_engineering_report.xlsx"), "output.filename")
    keep_plot_assets = output_raw.get("keep_plot_assets", True)
    if not isinstance(keep_plot_assets, bool):
        raise ConfigurationError("output.keep_plot_assets must be true or false")

    return ExcelReportConfig(
        version=version,
        title=title,
        subtitle=subtitle,
        report_sheet=report_sheet,
        metadata_sheet=metadata_sheet,
        channel_ids=channel_ids,
        channel_metadata=channel_metadata,
        top_rms_statistic_ids=top_rms,
        kpi_statistic_ids=kpis,
        bottom_operations=bottom_operations,
        bottom_summary_statistic_ids=bottom_summary,
        layout_profile=layout_profile,
        plot_placement=plot_placement,
        blank_separator_columns=blank_separator_columns,
        channel_width=channel_width,
        header_row_height=header_row_height,
        unit_row_height=unit_row_height,
        plot_ids=plot_ids,
        plot_columns=plot_columns,
        plot_width_px=plot_width_px,
        plot_height_px=plot_height_px,
        native_chart_ids=native_chart_ids,
        chart_layout=chart_layout,
        rms_merges=rms_merges,
        output_filename=output_filename,
        keep_plot_assets=keep_plot_assets,
    )


def generate_excel_report(
    input_file: str | Path,
    report_config_file: str | Path,
    statistics_config_file: str | Path,
    plotting_config_file: str | Path,
    output_dir: str | Path,
    import_options: ImportOptions | None = None,
    math_config_file: str | Path | None = None,
    *,
    precomputed_statistics_result: StatisticsResult | None = None,
    precomputed_plotting_result: PlottingResult | None = None,
) -> ExcelReportResult:
    """Generate a deterministic Excel engineering report from the validated processing layers."""

    config_path = Path(report_config_file).expanduser().resolve()
    config = load_excel_report_config(config_path)
    destination = Path(output_dir).expanduser().resolve()
    destination.mkdir(parents=True, exist_ok=True)
    generated_plot_assets = precomputed_plotting_result is None
    plot_assets_dir = destination / "plot_assets"

    if precomputed_statistics_result is None:
        statistics_result = calculate_statistics(
            input_file,
            statistics_config_file,
            import_options,
            math_config_file=math_config_file,
        )
    else:
        statistics_result = precomputed_statistics_result

    if precomputed_plotting_result is None:
        plot_assets_dir.mkdir(parents=True, exist_ok=True)
        plotting_result = render_plots(
            input_file,
            plotting_config_file,
            plot_assets_dir,
            import_options,
            math_config_file=math_config_file,
        )
    else:
        plotting_result = precomputed_plotting_result
        if plotting_result.rendered_plots:
            plot_assets_dir = Path(plotting_result.rendered_plots[0].output_file).expanduser().resolve().parent
        else:
            plot_assets_dir.mkdir(parents=True, exist_ok=True)

    channels_by_id = statistics_result.channels_by_id
    values_by_id = statistics_result.values_by_id
    missing_channels = [channel_id for channel_id in config.channel_ids if channel_id not in channels_by_id]
    if missing_channels:
        raise ExcelReportError("Configured report channel IDs were not found: " + ", ".join(missing_channels))
    report_channels = [_apply_channel_metadata(channels_by_id[channel_id], config.channel_metadata.get(channel_id)) for channel_id in config.channel_ids]

    statistics_by_id = {item.statistic_id: item for item in statistics_result.statistics}
    missing_statistics = sorted(
        set(config.top_rms_statistic_ids + config.kpi_statistic_ids + config.bottom_summary_statistic_ids) - set(statistics_by_id)
    )
    if missing_statistics:
        raise ExcelReportError(
            "Configured report statistic IDs were not found in the statistics result: "
            + ", ".join(missing_statistics)
        )

    if config.layout_profile == "sergio_reference" and config.native_chart_ids:
        invisible_rms = [
            statistic_id
            for statistic_id in config.top_rms_statistic_ids
            if statistics_by_id[statistic_id].channel_id not in config.channel_ids
        ]
        invisible_bottom = [
            statistic_id
            for statistic_id in config.bottom_summary_statistic_ids
            if statistics_by_id[statistic_id].channel_id not in config.channel_ids
        ]
        if invisible_rms:
            raise ExcelReportError(
                "Sergio-reference RMS statistics must target exported report channels: "
                + ", ".join(invisible_rms)
            )
        if invisible_bottom:
            raise ExcelReportError(
                "Sergio-reference bottom statistics must target exported report channels: "
                + ", ".join(invisible_bottom)
            )

    plots_by_id = {item.plot_id: item for item in plotting_result.rendered_plots}
    missing_plots = sorted(set(config.plot_ids) - set(plots_by_id))
    if missing_plots:
        raise ExcelReportError(
            "Configured report plot IDs were not found in the plotting result: " + ", ".join(missing_plots)
        )
    plotting_config = load_plotting_config(plotting_config_file)
    native_chart_definitions = {
        definition.plot_id: definition
        for definition in plotting_config.plots
        if definition.plot_id in config.native_chart_ids
    }
    missing_native_charts = sorted(set(config.native_chart_ids) - set(native_chart_definitions))
    if missing_native_charts:
        raise ExcelReportError(
            "Configured native Excel chart IDs were not found in the plotting configuration: "
            + ", ".join(missing_native_charts)
        )

    report_path = destination / config.output_filename
    manifest_path = destination / "excel_report_manifest.json"
    summary_path = destination / "excel_report_summary.txt"

    _write_workbook(
        report_path,
        config,
        input_file=Path(input_file).expanduser().resolve(),
        report_config_file=config_path,
        math_config_file=Path(math_config_file).expanduser().resolve() if math_config_file else None,
        statistics_config_file=Path(statistics_config_file).expanduser().resolve(),
        plotting_config_file=Path(plotting_config_file).expanduser().resolve(),
        statistics_result=statistics_result,
        report_channels=report_channels,
        values_by_id=values_by_id,
        statistics_by_id=statistics_by_id,
        plots_by_id=plots_by_id,
        native_chart_definitions=native_chart_definitions,
    )

    result = ExcelReportResult(
        report_path=report_path,
        manifest_path=manifest_path,
        summary_path=summary_path,
        plot_assets_dir=plot_assets_dir,
        config_path=config_path,
        config=config,
        statistics_result=statistics_result,
        plotting_result=plotting_result,
        report_channels=report_channels,
    )
    manifest_path.write_text(json.dumps(_manifest(result, statistics_config_file, plotting_config_file, math_config_file), indent=2), encoding="utf-8")
    summary_path.write_text(_summary(result), encoding="utf-8")

    if not config.keep_plot_assets and generated_plot_assets:
        shutil.rmtree(plot_assets_dir, ignore_errors=True)

    return result


def generate_profile_excel_report(
    input_file: str | Path,
    profile_file: str | Path,
    output_dir: str | Path,
    import_options: ImportOptions | None = None,
    *,
    output_filename: str | None = None,
    report_type: str | None = None,
    template_file: str | Path | None = None,
) -> ProfileExcelReportResult:
    """Generate a profile-driven RoboSprayer engineering workbook.

    This adapter uses the validated profile processing layers as numerical
    authority and writes semantic raw/math report channels, statistics, KPIs,
    and profile-rendered plot images into one workbook.
    """

    source_path = Path(input_file).expanduser().resolve()
    profile_path = Path(profile_file).expanduser().resolve()
    profile = load_reporting_profile(profile_path)
    dataset = load_data_file(source_path, import_options)
    resolution = resolve_profile(dataset, profile)
    if not resolution.is_valid:
        missing = [item.definition.semantic_name for item in resolution.missing_required]
        ambiguous = [item.definition.semantic_name for item in resolution.ambiguous]
        mismatches = [item.definition.semantic_name for item in resolution.unit_mismatches]
        raise ExcelReportError(
            "Profile raw channels are not fully resolved: "
            f"missing={missing}; ambiguous={ambiguous}; unit_mismatches={mismatches}"
        )

    math_result = calculate_profile_math_channels(dataset, profile, resolution)
    if math_result.unavailable_required:
        raise ExcelReportError(
            "Required profile MATH channels are unavailable: "
            + ", ".join(item.definition.semantic_name for item in math_result.unavailable_required)
        )

    statistics_result = calculate_profile_statistics(dataset, profile, resolution, math_result)
    if not statistics_result.is_complete:
        missing_stats = [item.definition.statistic_id for item in statistics_result.unavailable_required_statistics]
        missing_kpis = [item.definition.kpi_id for item in statistics_result.unavailable_required_kpis]
        raise ExcelReportError(
            "Required profile statistics/KPIs are unavailable: "
            f"statistics={missing_stats}; kpis={missing_kpis}"
        )

    destination = Path(output_dir).expanduser().resolve()
    destination.mkdir(parents=True, exist_ok=True)
    plot_assets_dir = destination / "profile_plot_assets"
    plotting_result = render_profile_plots(dataset, profile, plot_assets_dir, resolution, math_result)
    if plotting_result.unavailable_plots:
        raise ExcelReportError(
            "Profile plots are unavailable: "
            + ", ".join(item.definition.plot_id for item in plotting_result.unavailable_plots)
        )

    channels_by_name = plotting_result.channels_by_semantic_name
    report_channels = _profile_report_channels(
        profile,
        channels_by_name,
        template_file=_profile_template_file(profile, template_file),
    )
    comparison = _profile_template_comparison_rows(profile, dataset, report_channels, statistics_result, plotting_result)
    report_path = destination / _profile_output_filename(profile, output_filename)
    manifest_path = destination / "profile_excel_report_manifest.json"
    summary_path = destination / "profile_excel_report_summary.txt"
    report_type = report_type or (profile.metadata.powertrain or profile.profile_id)

    _write_profile_workbook(
        report_path,
        source_path=source_path,
        profile_path=profile_path,
        report_type=report_type,
        profile=profile,
        dataset=dataset,
        resolution=resolution,
        math_result=math_result,
        statistics_result=statistics_result,
        plotting_result=plotting_result,
        report_channels=report_channels,
        values_by_name=plotting_result.values_by_semantic_name,
        template_comparison=comparison,
    )

    result = ProfileExcelReportResult(
        report_path=report_path,
        manifest_path=manifest_path,
        summary_path=summary_path,
        plot_assets_dir=plot_assets_dir,
        dataset=dataset,
        profile=profile,
        resolution=resolution,
        math_result=math_result,
        statistics_result=statistics_result,
        plotting_result=plotting_result,
        report_channels=report_channels,
        template_comparison=comparison,
    )
    manifest_path.write_text(
        json.dumps(_profile_manifest(result, source_path, profile_path, report_type), indent=2),
        encoding="utf-8",
    )
    summary_path.write_text(_profile_summary(result), encoding="utf-8")
    return result


def _profile_output_filename(profile: ReportingProfile, output_filename: str | None) -> str:
    if output_filename is not None:
        return _plain_xlsx_filename(output_filename, "output_filename")
    return f"{profile.profile_id}_profile_report.xlsx"


def _profile_template_file(profile: ReportingProfile, template_file: str | Path | None) -> Path | None:
    if template_file is not None:
        path = Path(template_file).expanduser().resolve()
        if not path.exists():
            raise ExcelReportError(f"Profile Excel template file does not exist: {path}")
        return path
    if profile.profile_id != "robosprayer_electric":
        return None
    candidate = Path("reference_files/Robo_Sprayer_Electrification_Tamplate_Electric.xlsx").resolve()
    return candidate if candidate.exists() else None


def _profile_report_channels(
    profile: ReportingProfile,
    channels_by_name: Mapping[str, ChannelInfo],
    *,
    template_file: Path | None = None,
) -> list[ChannelInfo]:
    if template_file is not None:
        return _profile_report_channels_from_template(profile, channels_by_name, template_file)

    channels: list[ChannelInfo] = []
    for definition in profile.raw_channels:
        if definition.semantic_name not in channels_by_name:
            raise ExcelReportError(f"Resolved raw profile channel is unavailable: {definition.semantic_name}")
        channel = channels_by_name[definition.semantic_name]
        channels.append(
            replace(
                channel,
                channel_id=definition.semantic_name,
                source_name=definition.source_name,
                display_name=definition.report_name,
                unit=definition.unit,
                kind=definition.channel_type.lower(),
            )
        )
    for definition in profile.math_channels:
        if definition.semantic_name not in channels_by_name:
            raise ExcelReportError(f"Calculated profile MATH channel is unavailable: {definition.semantic_name}")
        channel = channels_by_name[definition.semantic_name]
        channels.append(
            replace(
                channel,
                channel_id=definition.semantic_name,
                source_name=definition.source_name,
                display_name=definition.report_name,
                unit=definition.unit,
                kind="math",
            )
        )
    return channels


def _profile_report_channels_from_template(
    profile: ReportingProfile,
    channels_by_name: Mapping[str, ChannelInfo],
    template_file: Path,
) -> list[ChannelInfo]:
    raw_by_name = profile.raw_by_semantic_name()
    math_by_name = profile.math_by_semantic_name()
    definitions: dict[str, Any] = {**raw_by_name, **math_by_name}
    definitions_by_exact_key: dict[tuple[str, str, str], list[str]] = {}
    definitions_by_key: dict[tuple[str, str], list[str]] = {}
    definitions_by_report_key: dict[tuple[str, str], list[str]] = {}
    for semantic_name, definition in definitions.items():
        channel_type = definition.channel_type.upper()
        definitions_by_exact_key.setdefault(
            (normalized_name(definition.source_name), normalized_name(definition.report_name), channel_type),
            [],
        ).append(semantic_name)
        definitions_by_key.setdefault((normalized_name(definition.source_name), channel_type), []).append(semantic_name)
        definitions_by_report_key.setdefault((normalized_name(definition.report_name), channel_type), []).append(semantic_name)

    channels: list[ChannelInfo] = []
    seen: set[str] = set()
    missing: list[str] = []
    for entry in _template_channel_order_entries(template_file):
        channel_type = entry["channel_type"].upper()
        semantic_name = _consume_template_match(
            definitions_by_exact_key.get(
                (normalized_name(entry["source_name"]), normalized_name(entry["report_name"]), channel_type),
                [],
            ),
            seen,
        )
        if semantic_name is None:
            semantic_name = _consume_template_match(
                definitions_by_key.get((normalized_name(entry["source_name"]), channel_type), []),
                seen,
            )
        if semantic_name is None:
            semantic_name = _consume_template_match(
                definitions_by_report_key.get((normalized_name(entry["report_name"]), channel_type), []),
                seen,
            )
        if semantic_name is None:
            missing.append(f"{entry['source_name']} / {entry['report_name']} [{channel_type}]")
            continue
        seen.add(semantic_name)
        definition = definitions[semantic_name]
        channel = channels_by_name[semantic_name]
        kind = "math" if channel_type == "MATH" else definition.channel_type.lower()
        channels.append(
            replace(
                channel,
                channel_id=semantic_name,
                source_name=definition.source_name,
                display_name=definition.report_name,
                unit=definition.unit,
                kind=kind,
            )
        )

    expected = set(definitions)
    omitted = sorted(expected - seen)
    if missing or omitted:
        raise ExcelReportError(
            "Sergio template channel order does not match the active profile: "
            f"missing_template_matches={missing[:10]}; omitted_profile_channels={omitted[:10]}"
        )
    return channels


def _consume_template_match(candidates: list[str], seen: set[str]) -> str | None:
    for semantic_name in candidates:
        if semantic_name not in seen:
            return semantic_name
    return None


def _template_channel_order_entries(template_file: Path) -> list[dict[str, str]]:
    workbook = load_workbook(template_file, read_only=True, data_only=True)
    try:
        sheet = next(
            worksheet
            for worksheet in workbook.worksheets
            if worksheet.cell(2, 2).value == "VSM Name"
            and worksheet.cell(2, 3).value == "Astauto Name"
            and worksheet.cell(2, 4).value == "Channel Type"
        )
    except StopIteration as exc:
        raise ExcelReportError(f"Template workbook has no channel selection sheet: {template_file}") from exc

    entries: list[dict[str, str]] = []
    for row in range(3, sheet.max_row + 1):
        source_name = sheet.cell(row, 2).value
        report_name = sheet.cell(row, 3).value
        channel_type = sheet.cell(row, 4).value
        if source_name is None and report_name is None and channel_type is None:
            continue
        if not isinstance(source_name, str) or not isinstance(report_name, str) or not isinstance(channel_type, str):
            raise ExcelReportError(f"Invalid template channel-selection row {row} in {template_file}")
        entries.append(
            {
                "source_name": source_name.strip(),
                "report_name": report_name.strip(),
                "channel_type": channel_type.strip(),
                "for_plot": "yes" if sheet.cell(row, 5).value else "",
            }
        )
    return entries


def _write_profile_workbook(
    output_path: Path,
    *,
    source_path: Path,
    profile_path: Path,
    report_type: str,
    profile: ReportingProfile,
    dataset: Any,
    resolution: ProfileResolutionResult,
    math_result: ProfileMathResult,
    statistics_result: ProfileStatisticsResult,
    plotting_result: ProfilePlottingResult,
    report_channels: list[ChannelInfo],
    values_by_name: Mapping[str, Any],
    template_comparison: list[dict[str, str]],
) -> None:
    workbook = Workbook()
    report = workbook.active
    report.title = _profile_sheet_name(profile)
    mapping_sheet = workbook.create_sheet("Rename From VSM to Astauto")
    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.sheet_state = "hidden"

    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="143642")
    vsm_fill = PatternFill("solid", fgColor="1F4E78")
    avl_fill = PatternFill("solid", fgColor="375623")
    math_fill = PatternFill("solid", fgColor="C65911")
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    kpi_fill = PatternFill("solid", fgColor="548235")
    white_bold = Font(color="FFFFFF", bold=True)

    _write_profile_report_sheet(
        report,
        profile=profile,
        source_path=source_path,
        report_type=report_type,
        statistics_result=statistics_result,
        plotting_result=plotting_result,
        report_channels=report_channels,
        values_by_name=values_by_name,
        border=border,
        title_fill=title_fill,
        vsm_fill=vsm_fill,
        avl_fill=avl_fill,
        math_fill=math_fill,
        label_fill=label_fill,
        kpi_fill=kpi_fill,
        white_bold=white_bold,
    )
    _write_profile_channel_mapping_sheet(
        mapping_sheet,
        profile=profile,
        report_channels=report_channels,
        border=border,
        label_fill=label_fill,
        white_bold=white_bold,
    )
    _write_profile_metadata_sheet(
        metadata_sheet,
        source_path=source_path,
        profile_path=profile_path,
        profile=profile,
        report_type=report_type,
        dataset=dataset,
        resolution=resolution,
        math_result=math_result,
        statistics_result=statistics_result,
        plotting_result=plotting_result,
        report_channels=report_channels,
        border=border,
        label_fill=label_fill,
        white_bold=white_bold,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_profile_report_sheet(
    sheet: Any,
    *,
    profile: ReportingProfile,
    source_path: Path,
    report_type: str,
    statistics_result: ProfileStatisticsResult,
    plotting_result: ProfilePlottingResult,
    report_channels: list[ChannelInfo],
    values_by_name: Mapping[str, Any],
    border: Border,
    title_fill: PatternFill,
    vsm_fill: PatternFill,
    avl_fill: PatternFill,
    math_fill: PatternFill,
    label_fill: PatternFill,
    kpi_fill: PatternFill,
    white_bold: Font,
) -> None:
    channel_count = len(report_channels)
    sample_count = statistics_result.dataset.quality.sample_count
    data_start_row = _PROFILE_DATA_START_ROW
    data_end_row = data_start_row + sample_count - 1
    channel_columns = {channel.channel_id: index for index, channel in enumerate(report_channels, start=1)}

    summary_start_col = channel_count + 2

    title = sheet.cell(1, 1, profile.metadata.name)
    title.fill = title_fill
    title.font = Font(color="FFFFFF", bold=True, size=13)
    title.border = border
    sheet.cell(2, 1, f"{client_display_filename(source_path)} | {sample_count} samples | {report_type}").border = border

    _write_profile_rms_blocks(
        sheet,
        statistics_result,
        channel_columns,
        title_fill=title_fill,
        border=border,
        white_bold=white_bold,
    )
    _write_profile_right_summary(
        sheet,
        statistics_result,
        start_col=summary_start_col,
        border=border,
        label_fill=label_fill,
        white_bold=white_bold,
    )

    for col_index, channel in enumerate(report_channels, start=1):
        channel_type = channel.kind.upper()
        fill = math_fill if channel_type == "MATH" else avl_fill if channel_type == "AVL" else vsm_fill
        name_cell = sheet.cell(3, col_index, _main_report_header(channel.display_name, channel.unit))
        unit_cell = sheet.cell(4, col_index, channel.unit or "-")
        for cell in (name_cell, unit_cell):
            cell.fill = fill
            cell.font = white_bold
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_index, channel in enumerate(report_channels, start=1):
        values = np.asarray(values_by_name[channel.channel_id], dtype=np.float64)
        if values.size != sample_count:
            raise ExcelReportError(
                f"Profile channel '{channel.channel_id}' has {values.size} samples; expected {sample_count}"
            )
        number_format = _number_format(channel.unit)
        for row_offset, raw_value in enumerate(values):
            cell = sheet.cell(data_start_row + row_offset, col_index, float(raw_value))
            cell.number_format = number_format

    operation_rows = _bottom_statistic_rows(statistics_result)
    for operation, row in operation_rows.items():
        label = sheet.cell(row, 1, operation.upper())
        label.fill = label_fill
        label.font = Font(bold=True)
        label.border = border
    for item in statistics_result.statistics:
        if item.definition.placement_group == "top_rms" or item.definition.operation not in operation_rows:
            continue
        col = channel_columns.get(item.target_channel)
        if col is None:
            continue
        cell = sheet.cell(operation_rows[item.definition.operation], col, item.value)
        cell.number_format = "0.000000"
        cell.fill = label_fill
        cell.border = border
        cell.font = Font(bold=True)

    _write_profile_plots_on_report_sheet(
        sheet,
        plotting_result,
        start_row=6,
        start_col=summary_start_col,
        border=border,
        label_fill=label_fill,
        white_bold=white_bold,
    )

    sheet.freeze_panes = f"B{data_start_row}"
    for col_index, channel in enumerate(report_channels, start=1):
        width = min(max(len(channel.display_name), 9), 16 if channel.kind != "math" else 18)
        sheet.column_dimensions[get_column_letter(col_index)].width = width
    for col_index in range(summary_start_col, summary_start_col + 42):
        sheet.column_dimensions[get_column_letter(col_index)].width = 13
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 42
    sheet.row_dimensions[4].height = 22


def _top_summary_column(
    target_channel: str,
    report_channels: list[ChannelInfo],
    channel_columns: Mapping[str, int],
) -> int:
    if target_channel in channel_columns:
        return channel_columns[target_channel]
    target_prefix = normalized_name(target_channel)
    for channel in report_channels:
        if normalized_name(channel.channel_id).startswith(target_prefix):
            return channel_columns[channel.channel_id]
    return min(9, len(report_channels))


def _write_profile_rms_blocks(
    sheet: Any,
    statistics_result: ProfileStatisticsResult,
    channel_columns: Mapping[str, int],
    *,
    title_fill: PatternFill,
    border: Border,
    white_bold: Font,
) -> None:
    occupied: list[tuple[int, int]] = []
    for item in statistics_result.statistics:
        if item.definition.placement_group != "top_rms":
            continue
        start_col, end_col, value_col = _rms_block_columns(item.definition.statistic_id, item.target_channel, channel_columns)
        start_col, end_col, value_col = _non_overlapping_rms_range(start_col, end_col, value_col, occupied)
        occupied.append((start_col, end_col))
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        label = sheet.cell(1, start_col, item.definition.display_name or item.channel_display_name)
        value = sheet.cell(2, value_col, _excel_display_number(item.value))
        for col in range(start_col, end_col + 1):
            for row in (1, 2):
                cell = sheet.cell(row, col)
                cell.fill = title_fill
                cell.font = white_bold
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value.number_format = "0.000000"


def _non_overlapping_rms_range(
    start_col: int,
    end_col: int,
    value_col: int,
    occupied: list[tuple[int, int]],
) -> tuple[int, int, int]:
    width = end_col - start_col
    value_offset = value_col - start_col
    for occupied_start, occupied_end in occupied:
        if start_col <= occupied_end and end_col >= occupied_start:
            start_col = occupied_end + 1
            end_col = start_col + width
            value_col = start_col + value_offset
    return start_col, end_col, value_col


def _rms_block_columns(
    statistic_id: str,
    target_channel: str,
    channel_columns: Mapping[str, int],
) -> tuple[int, int, int]:
    if statistic_id == "battery_power_rms" and "total_edu_elect_power" in channel_columns:
        start_col = channel_columns["total_edu_elect_power"]
        return start_col, start_col + 2, start_col + 2
    if statistic_id == "battery_heatflow_rms" and "battery_heatflow_squared" in channel_columns:
        start_col = channel_columns["battery_heatflow_squared"]
        return start_col, start_col + 3, start_col
    target_col = channel_columns.get(target_channel, min(channel_columns.values()))
    return max(1, target_col - 1), target_col + 1, target_col


def _write_profile_right_summary(
    sheet: Any,
    statistics_result: ProfileStatisticsResult,
    *,
    start_col: int,
    border: Border,
    label_fill: PatternFill,
    white_bold: Font,
) -> None:
    for offset, item in enumerate(_profile_right_summary_items(statistics_result)):
        col = start_col + offset
        label = sheet.cell(3, col, item["label"])
        value = sheet.cell(4, col, _excel_display_number(item["value"]))
        label.fill = label_fill
        label.font = white_bold
        label.border = border
        label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value.border = border
        value.alignment = Alignment(horizontal="center", vertical="center")
        value.number_format = "0.000000"


def _profile_right_summary_items(statistics_result: ProfileStatisticsResult) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for statistic in statistics_result.statistics:
        label = statistic.definition.display_name or statistic.channel_display_name
        unit = statistic.channel_unit or statistic.definition.unit
        items.append({"label": _summary_label(label, unit), "value": statistic.value})
    for kpi in statistics_result.kpis:
        label = _client_kpi_label(kpi.definition.kpi_id, kpi.definition.display_name)
        items.append({"label": _summary_label(label, kpi.definition.unit), "value": kpi.value})
    return items


def _summary_label(label: str, unit: str | None) -> str:
    return f"{label} [{unit}]" if unit and unit != "-" else label


def _main_report_header(display_name: str, unit: str | None) -> str:
    suffix = f" ({unit})" if unit else ""
    if suffix and display_name.endswith(suffix):
        return display_name[: -len(suffix)]
    return display_name


def _excel_display_number(value: Any) -> Any:
    if isinstance(value, (int, float)) and abs(float(value)) < 1e-12:
        return 0.0
    return value


def _bottom_statistic_rows(statistics_result: ProfileStatisticsResult) -> dict[str, int]:
    operations = ["max", "min", "last", "first", "sum"]
    present = {
        item.definition.operation
        for item in statistics_result.statistics
        if item.definition.placement_group != "top_rms"
    }
    return {
        operation: _PROFILE_DATA_START_ROW + statistics_result.dataset.quality.sample_count + index
        for index, operation in enumerate(operations)
        if operation in present
    }


def _write_profile_kpi_block(
    sheet: Any,
    statistics_result: ProfileStatisticsResult,
    start_row: int,
    *,
    border: Border,
    kpi_fill: PatternFill,
    white_bold: Font,
) -> None:
    sheet.cell(start_row, 1, "Engineering KPIs")
    sheet.cell(start_row, 1).fill = kpi_fill
    sheet.cell(start_row, 1).font = white_bold
    sheet.cell(start_row, 1).border = border
    headers = ("KPI", "Value", "Unit")
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row + 1, col, header)
        cell.fill = kpi_fill
        cell.font = white_bold
        cell.border = border
    for row_offset, item in enumerate(statistics_result.kpis, start=2):
        row = start_row + row_offset
        values = (
            _client_kpi_label(item.definition.kpi_id, item.definition.display_name),
            item.value,
            item.definition.unit or "-",
        )
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col, value)
            cell.border = border
            if col == 2:
                cell.number_format = "0.000000"


def _write_profile_plots_on_report_sheet(
    sheet: Any,
    plotting_result: ProfilePlottingResult,
    start_row: int,
    start_col: int,
    *,
    border: Border,
    label_fill: PatternFill,
    white_bold: Font,
) -> None:
    rendered = _ordered_profile_plots(plotting_result.rendered_plots)
    for index, plot in enumerate(rendered):
        row = start_row + (index // 6) * 17
        col = start_col + (index % 6) * 6
        title = _client_plot_title(plot.plot_id, plot.title)
        label = sheet.cell(row, col, title)
        label.fill = label_fill
        label.font = Font(bold=True)
        label.border = border
        label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        image = XLImage(plot.png_file)
        image.width = 300
        image.height = 180
        sheet.add_image(image, f"{get_column_letter(col)}{row + 1}")


def _ordered_profile_plots(rendered_plots: list[Any]) -> list[Any]:
    preferred = [
        "speed_vs_distance",
        "speed_vs_time",
        "battery_energy_distance_based",
        "battery_energy_time_based",
        "power_at_wheels_and_edu",
        "battery_soc",
        "battery_power_charge_discharge",
        "energy_released",
        "auxiliaries_energy_consumption",
        "tyres_energy_consumption",
        "agrochemical_discharge_vs_distance",
        "agrochemical_discharge_and_charge_vs_time",
    ]
    by_id = {plot.plot_id: plot for plot in rendered_plots}
    ordered = [by_id[plot_id] for plot_id in preferred if plot_id in by_id]
    ordered.extend(plot for plot in rendered_plots if plot.plot_id not in set(preferred))
    return ordered


def _write_profile_channel_mapping_sheet(
    sheet: Any,
    *,
    profile: ReportingProfile,
    report_channels: list[ChannelInfo],
    border: Border,
    label_fill: PatternFill,
    white_bold: Font,
) -> None:
    raw_by_name = profile.raw_by_semantic_name()
    math_by_name = profile.math_by_semantic_name()
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    sheet.cell(1, 1, f"{profile.metadata.name} - Rename From VSM to Astauto")
    sheet.cell(1, 1).fill = label_fill
    sheet.cell(1, 1).font = white_bold
    headers = [
        "VSM Name",
        "Astauto Name",
        "Channel Type",
        "For Plot",
        "Unit",
        "Semantic ID",
        "Order",
        "Formula / dependencies",
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(2, col, header)
        cell.fill = label_fill
        cell.font = Font(bold=True)
        cell.border = border
    for row, channel in enumerate(report_channels, start=3):
        raw_definition = raw_by_name.get(channel.channel_id)
        math_definition = math_by_name.get(channel.channel_id)
        definition = raw_definition or math_definition
        for_plot = "yes" if definition and definition.for_plot else ""
        formula = ""
        if math_definition is not None:
            formula = math_definition.formula or math_definition.expression or ""
            if formula:
                formula = f"Formula: {formula}"
            if math_definition.dependencies:
                formula = f"{formula} | dependencies: {', '.join(math_definition.dependencies)}"
        values = [
            definition.source_name if definition else channel.source_name,
            definition.report_name if definition else channel.display_name,
            definition.channel_type if definition else channel.kind.upper(),
            for_plot,
            definition.unit or "-" if definition else channel.unit or "-",
            channel.channel_id,
            row - 2,
            formula,
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A3"
    for col, width in enumerate((42, 32, 14, 10, 12, 36, 8, 72), start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width


def _client_kpi_label(kpi_id: str, display_name: str | None) -> str:
    if kpi_id == "range_85_battery_km":
        return "Range for 85% Battery"
    return display_name or kpi_id.replace("_", " ").title()


def _client_plot_title(plot_id: str, title: str) -> str:
    if plot_id == "agrochemical_discharge_and_charge_vs_time":
        return "Agrochemical Discharge and Battery SOC Vs Time"
    return title


def _write_profile_statistics_sheet(
    sheet: Any,
    statistics_result: ProfileStatisticsResult,
    *,
    border: Border,
    label_fill: PatternFill,
    kpi_fill: PatternFill,
    white_bold: Font,
) -> None:
    headers = [
        "Statistic ID",
        "Target",
        "Operation",
        "Display name",
        "Value",
        "Unit",
        "Placement group",
        "Sample count",
        "Used samples",
        "Omitted samples",
    ]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col, header)
        cell.fill = label_fill
        cell.font = Font(bold=True)
        cell.border = border
    for row, item in enumerate(statistics_result.statistics, start=2):
        values = [
            item.definition.statistic_id,
            item.target_channel,
            item.definition.operation,
            item.definition.display_name or item.channel_display_name,
            item.value,
            item.channel_unit or "-",
            item.definition.placement_group or "-",
            item.sample_count,
            item.used_sample_count,
            item.omitted_sample_count,
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col, value)
            cell.border = border
            if col == 5:
                cell.number_format = "0.000000"

    kpi_start = len(statistics_result.statistics) + 4
    kpi_headers = ["KPI ID", "Display name", "Value", "Unit", "Dependencies", "Placement group"]
    for col, header in enumerate(kpi_headers, start=1):
        cell = sheet.cell(kpi_start, col, header)
        cell.fill = kpi_fill
        cell.font = white_bold
        cell.border = border
    for row, item in enumerate(statistics_result.kpis, start=kpi_start + 1):
        values = [
            item.definition.kpi_id,
            item.definition.display_name or item.definition.kpi_id,
            item.value,
            item.definition.unit or "-",
            ", ".join(item.definition.dependencies),
            item.definition.placement_group or "-",
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col, value)
            cell.border = border
            if col == 3:
                cell.number_format = "0.000000"

    sheet.freeze_panes = "A2"
    for col in range(1, 11):
        sheet.column_dimensions[get_column_letter(col)].width = 22


def _write_profile_plots_sheet(
    sheet: Any,
    plotting_result: ProfilePlottingResult,
    *,
    border: Border,
    label_fill: PatternFill,
    white_bold: Font,
) -> None:
    headers = ["Plot ID", "Title", "Reference chart", "X channel", "Primary series", "Secondary series", "PNG file"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col, header)
        cell.fill = label_fill
        cell.font = Font(bold=True)
        cell.border = border
    for row, plot in enumerate(plotting_result.rendered_plots, start=2):
        values = [
            plot.plot_id,
            plot.title,
            plot.reference_chart_number or "-",
            plot.x_channel_id,
            ", ".join(plot.primary_series_ids),
            ", ".join(plot.secondary_series_ids),
            plot.png_file,
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row, col, value)
            cell.border = border

    image_start_row = len(plotting_result.rendered_plots) + 4
    for index, plot in enumerate(plotting_result.rendered_plots):
        row = image_start_row + (index // 2) * 20
        col = 1 + (index % 2) * 6
        label = sheet.cell(row, col, plot.title)
        label.fill = label_fill
        label.font = white_bold
        label.border = border
        image = XLImage(plot.png_file)
        image.width = 480
        image.height = 288
        sheet.add_image(image, f"{get_column_letter(col)}{row + 1}")

    for col in range(1, 13):
        sheet.column_dimensions[get_column_letter(col)].width = 20


def _write_profile_metadata_sheet(
    sheet: Any,
    *,
    source_path: Path,
    profile_path: Path,
    profile: ReportingProfile,
    report_type: str,
    dataset: Any,
    resolution: ProfileResolutionResult,
    math_result: ProfileMathResult,
    statistics_result: ProfileStatisticsResult,
    plotting_result: ProfilePlottingResult,
    report_channels: list[ChannelInfo],
    border: Border,
    label_fill: PatternFill,
    white_bold: Font,
) -> None:
    raw_counts = _profile_raw_type_counts(profile)
    metadata = [
        ("Software version", __version__),
        ("Generated UTC", datetime.now(timezone.utc).isoformat()),
        ("Source file", client_display_filename(source_path)),
        ("Source display filename", client_display_filename(source_path)),
        ("Source SHA-256", dataset.quality.source_sha256),
        ("Source format", dataset.quality.file_type),
        ("Source raw channel count", dataset.quality.raw_channel_count),
        ("Source total channel count", dataset.quality.channel_count),
        ("Source sample count", dataset.quality.sample_count),
        ("Source data start row", dataset.quality.data_start_row),
        ("Source data end row", dataset.quality.data_end_row),
        ("Workbook data start row", _PROFILE_DATA_START_ROW),
        ("Workbook data end row", _PROFILE_DATA_START_ROW + dataset.quality.sample_count - 1),
        ("Profile file", _client_relative_path(profile_path)),
        ("Profile SHA-256", sha256_file(profile_path)),
        ("Profile ID", profile.profile_id),
        ("Profile name", profile.metadata.name),
        ("Powertrain", profile.metadata.powertrain or "-"),
        ("Report type", report_type),
        ("Resolved raw profile channels", len(resolution.resolved)),
        ("Exported report channels", len(report_channels)),
        ("Exported VSM raw channels", raw_counts["VSM"]),
        ("Exported AVL raw channels", raw_counts["AVL"]),
        ("Exported MATH channels", len(math_result.calculated_channels)),
        ("Configured statistics", statistics_result.configured_statistic_count),
        ("Calculated statistics", statistics_result.calculated_statistic_count),
        ("Configured KPIs", statistics_result.configured_kpi_count),
        ("Calculated KPIs", statistics_result.calculated_kpi_count),
        ("Configured plots", plotting_result.configured_plot_count),
        ("Rendered plots", plotting_result.rendered_plot_count),
        ("Plot series", plotting_result.series_count),
        ("Visible sheets", f"{_profile_sheet_name(profile)}, Rename From VSM to Astauto"),
        ("Hidden sheets", "Metadata"),
    ]
    sheet.cell(1, 1, "Field").fill = label_fill
    sheet.cell(1, 2, "Value").fill = label_fill
    for cell in (sheet.cell(1, 1), sheet.cell(1, 2)):
        cell.font = white_bold
        cell.border = border
    for row, (key, value) in enumerate(metadata, start=2):
        sheet.cell(row, 1, key).border = border
        sheet.cell(row, 2, value).border = border
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 96


def _write_profile_template_comparison_sheet(
    sheet: Any,
    rows: list[dict[str, str]],
    *,
    border: Border,
    label_fill: PatternFill,
    warning_fill: PatternFill,
    white_bold: Font,
) -> None:
    headers = ["Feature", "Sergio template", "Generated report", "Status", "Comments"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(1, col, header)
        cell.fill = label_fill
        cell.font = white_bold
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        values = [row["feature"], row["sergio_template"], row["generated_report"], row["status"], row["comments"]]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, col, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row["status"] == "INTENTIONAL CORRECTION":
                cell.fill = warning_fill
    for col, width in enumerate((28, 36, 36, 24, 64), start=1):
        sheet.column_dimensions[get_column_letter(col)].width = width


def _profile_template_comparison_rows(
    profile: ReportingProfile,
    dataset: Any,
    report_channels: list[ChannelInfo],
    statistics_result: ProfileStatisticsResult,
    plotting_result: ProfilePlottingResult,
) -> list[dict[str, str]]:
    raw_counts = _profile_raw_type_counts(profile)
    return [
        _comparison_row(
            "Channel selection",
            "Selected RoboSprayer Electric engineering channels",
            f"{len(profile.raw_channels)} raw + {len(profile.math_channels)} MATH channels",
            "PASS",
            "Selection is profile-owned and independent of runtime column IDs.",
        ),
        _comparison_row(
            "Channel order",
            "317-entry Electric selection table with interspersed MATH channels",
            f"{len(report_channels)} semantic report columns in Sergio/template order",
            "PASS",
            "Runtime column IDs remain internal; the visible order follows source/report/type template rows.",
        ),
        _comparison_row(
            "Channel names and units",
            "Sergio visible report labels and units",
            "Report labels and units come from the active YAML profile",
            "PASS",
            "Runtime IDs remain internal provenance only.",
        ),
        _comparison_row(
            "Sample geometry",
            "Electric source cycle length",
            f"{dataset.quality.sample_count} samples from imported data",
            "PASS",
            "Workbook row count is generated dynamically from source data.",
        ),
        _comparison_row(
            "Powertrain split",
            "VSM, AVL, and MATH regions",
            f"{raw_counts['VSM']} VSM, {raw_counts['AVL']} AVL, {len(profile.math_channels)} MATH",
            "PASS",
            "Channel type is color-coded on row 3/4 and documented on Rename From VSM to Astauto.",
        ),
        _comparison_row(
            "RMS calculations",
            "Battery power and heatflow RMS in top region",
            f"{statistics_result.calculated_statistic_count} profile statistics calculated in Python",
            "INTENTIONAL CORRECTION",
            "RMS values use current Electric sample count, avoiding stale Caiman denominators.",
        ),
        _comparison_row(
            "Statistics and KPIs",
            "Main-sheet bottom statistics and engineering KPI summary",
            f"{statistics_result.calculated_statistic_count} statistics + {statistics_result.calculated_kpi_count} KPIs on main report",
            "PASS",
            "Values are written as numbers, not stale Excel formulas.",
        ),
        _comparison_row(
            "Agrochemical channel",
            "Electric report agrochemical discharge/charge plots",
            "Profile semantic agrochemical channels retained",
            "PASS",
            "Inactive Electric values remain valid zero series and plotted accordingly.",
        ),
        _comparison_row(
            "Plots",
            "Sergio reference plot set",
            f"{plotting_result.rendered_plot_count} of {plotting_result.configured_plot_count} profile plots embedded on main sheet",
            "PASS",
            "The workbook embeds profile-rendered PNG plot assets.",
        ),
        _comparison_row(
            "Sergio template fidelity",
            "Report sheet plus channel-selection sheet",
            "Production workbook has main report, Rename From VSM to Astauto, and hidden Metadata",
            "PASS",
            "Development-only QA comparison content is omitted from visible client sheets.",
        ),
    ]


def _comparison_row(feature: str, sergio: str, generated: str, status: str, comments: str) -> dict[str, str]:
    return {
        "feature": feature,
        "sergio_template": sergio,
        "generated_report": generated,
        "status": status,
        "comments": comments,
    }


def _profile_manifest(
    result: ProfileExcelReportResult,
    source_path: Path,
    profile_path: Path,
    report_type: str,
) -> dict[str, Any]:
    return {
        "software_version": __version__,
        "report_file": str(result.report_path),
        "source_file": str(source_path),
        "source_sha256": result.dataset.quality.source_sha256,
        "profile_file": str(profile_path),
        "profile_sha256": sha256_file(profile_path),
        "profile_id": result.profile.profile_id,
        "profile_name": result.profile.metadata.name,
        "report_type": report_type,
        "source_sample_count": result.sample_count,
        "source_raw_channel_count": result.source_raw_channel_count,
        "exported_report_channel_count": result.report_channel_count,
        "exported_vsm_channel_count": result.vsm_count,
        "exported_avl_channel_count": result.avl_count,
        "exported_math_channel_count": result.math_count,
        "configured_statistic_count": result.statistics_result.configured_statistic_count,
        "calculated_statistic_count": result.statistic_count,
        "configured_kpi_count": result.statistics_result.configured_kpi_count,
        "calculated_kpi_count": result.kpi_count,
        "configured_plot_count": result.plotting_result.configured_plot_count,
        "rendered_plot_count": result.plot_count,
        "plot_series_count": result.plotting_result.series_count,
        "report_channel_ids": [channel.channel_id for channel in result.report_channels],
        "plot_ids": [plot.plot_id for plot in result.plotting_result.rendered_plots],
        "visible_sheet_names": [_profile_sheet_name(result.profile), "Rename From VSM to Astauto"],
        "hidden_sheet_names": ["Metadata"],
    }


def _profile_summary(result: ProfileExcelReportResult) -> str:
    stats = {item.definition.statistic_id: item for item in result.statistics_result.statistics}
    kpis = {item.definition.kpi_id: item for item in result.statistics_result.kpis}
    lines = [
        "PROFILE-DRIVEN EXCEL REPORT SUMMARY",
        "===================================",
        "Status: PASS",
        f"Profile: {result.profile.profile_id}",
        f"Source samples: {result.sample_count}",
        f"Source raw channels: {result.source_raw_channel_count}",
        f"Exported report channels: {result.report_channel_count}",
        f"VSM channels: {result.vsm_count}",
        f"AVL channels: {result.avl_count}",
        f"MATH channels: {result.math_count}",
        f"Statistics: {result.statistic_count}",
        f"KPIs: {result.kpi_count}",
        f"Plots: {result.plot_count}",
        f"Workbook: {result.report_path}",
        "",
        "Numerical regression:",
    ]
    for statistic_id in (
        "time_minutes_last",
        "distance_km_last",
        "battery_soc_last",
        "battery_power_rms",
        "battery_heatflow_rms",
        "agrochemical_discharge_max",
    ):
        if statistic_id in stats:
            item = stats[statistic_id]
            lines.append(f"  - {statistic_id}: {item.value:.12g} [{item.channel_unit or '-'}]")
    for kpi_id in ("battery_capacity_used", "battery_energy_consumption_wh_per_km", "range_85_battery_km"):
        if kpi_id in kpis:
            item = kpis[kpi_id]
            lines.append(f"  - {kpi_id}: {item.value:.12g} [{item.definition.unit or '-'}]")
    return "\n".join(lines) + "\n"


def _profile_sheet_name(profile: ReportingProfile) -> str:
    name = profile.metadata.name or profile.profile_id
    if len(name) <= 31 and not any(character in name for character in "[]:*?/\\"):
        return name
    return _sheet_name(name[:31].rstrip(), "profile report sheet")


def _client_relative_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(Path.cwd().resolve()))
    except ValueError:
        return client_display_filename(path)


def _profile_raw_type_counts(profile: ReportingProfile) -> dict[str, int]:
    counts = {"VSM": 0, "AVL": 0}
    for channel in profile.raw_channels:
        channel_type = channel.channel_type.upper()
        if channel_type in counts:
            counts[channel_type] += 1
    return counts


_PROFILE_DATA_START_ROW = 5


def _write_workbook(
    output_path: Path,
    config: ExcelReportConfig,
    *,
    input_file: Path,
    report_config_file: Path,
    math_config_file: Path | None,
    statistics_config_file: Path,
    plotting_config_file: Path,
    statistics_result: StatisticsResult,
    report_channels: list[ChannelInfo],
    values_by_id: Mapping[str, Any],
    statistics_by_id: Mapping[str, StatisticResult],
    plots_by_id: Mapping[str, Any],
    native_chart_definitions: Mapping[str, Any],
) -> None:
    if config.layout_profile == "sergio_reference":
        _write_sergio_reference_workbook(
            output_path,
            config,
            input_file=input_file,
            report_config_file=report_config_file,
            math_config_file=math_config_file,
            statistics_config_file=statistics_config_file,
            plotting_config_file=plotting_config_file,
            statistics_result=statistics_result,
            report_channels=report_channels,
            values_by_id=values_by_id,
            statistics_by_id=statistics_by_id,
            plots_by_id=plots_by_id,
            native_chart_definitions=native_chart_definitions,
        )
        return

    workbook = Workbook()
    report = workbook.active
    report.title = config.report_sheet
    metadata = workbook.create_sheet(config.metadata_sheet)

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    raw_fill = PatternFill("solid", fgColor="1F4E78")
    math_fill = PatternFill("solid", fgColor="C65911")
    title_fill = PatternFill("solid", fgColor="17365D")
    rms_fill = PatternFill("solid", fgColor="7030A0")
    kpi_fill = PatternFill("solid", fgColor="548235")
    stat_fill = PatternFill("solid", fgColor="D9EAF7")
    metadata_fill = PatternFill("solid", fgColor="5B9BD5")
    white_bold = Font(color="FFFFFF", bold=True)

    channel_count = len(report_channels)
    last_channel_col = get_column_letter(channel_count)
    report.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(4, channel_count))
    report.cell(1, 1, config.title)
    report.cell(1, 1).fill = title_fill
    report.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=14)
    report.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    report.merge_cells(start_row=2, start_column=1, end_row=2, end_column=min(4, channel_count))
    report.cell(2, 1, config.subtitle or input_file.name)
    report.cell(2, 1).font = Font(italic=True, color="666666")

    # Top RMS strip starts after the title block and may continue beyond data columns if needed.
    rms_start_col = 5
    for index, statistic_id in enumerate(config.top_rms_statistic_ids):
        item = statistics_by_id[statistic_id]
        col = rms_start_col + index * 2
        report.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        report.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        report.cell(1, col, item.display_name)
        report.cell(2, col, item.value)
        for row in (1, 2):
            cell = report.cell(row, col)
            cell.fill = rms_fill
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        report.cell(2, col).number_format = "0.000000"

    # Two-row channel header: name, then explicit type + unit so math/raw is not color-only.
    for col_index, channel in enumerate(report_channels, start=1):
        name_cell = report.cell(3, col_index, channel.display_name)
        meta_cell = report.cell(4, col_index, f"{channel.kind.upper()} | {channel.unit or '-'}")
        fill = math_fill if channel.kind == "math" else raw_fill
        for cell in (name_cell, meta_cell):
            cell.fill = fill
            cell.font = white_bold
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start_row = 5
    sample_count = statistics_result.sample_count
    for col_index, channel in enumerate(report_channels, start=1):
        values = values_by_id[channel.channel_id]
        for row_offset, value in enumerate(values, start=0):
            cell = report.cell(data_start_row + row_offset, col_index, float(value))
            cell.number_format = _number_format(channel.unit)

    data_end_row = data_start_row + sample_count - 1
    bottom_start_row = data_end_row + 2
    bottom_statistics = [
        item
        for item in statistics_result.statistics
        if item.operation in config.bottom_operations and item.channel_id in config.channel_ids
    ]
    stats_by_operation_channel = {(item.operation, item.channel_id): item for item in bottom_statistics}
    for offset, operation in enumerate(config.bottom_operations):
        row = bottom_start_row + offset
        label = report.cell(row, 1, operation.replace("_", " ").upper())
        label.fill = stat_fill
        label.font = Font(bold=True)
        label.border = border
        for col_index, channel in enumerate(report_channels, start=1):
            item = stats_by_operation_channel.get((operation, channel.channel_id))
            if item is None:
                continue
            cell = report.cell(row, col_index, item.value)
            cell.fill = stat_fill
            cell.font = Font(bold=True)
            cell.border = border
            cell.number_format = _number_format(item.channel_unit)

    # KPI strip is placed to the right of the selected data channels, like the reference workbook.
    kpi_start_col = channel_count + 2
    for index, statistic_id in enumerate(config.kpi_statistic_ids):
        item = statistics_by_id[statistic_id]
        col = kpi_start_col + index
        header = report.cell(3, col, item.display_name)
        value = report.cell(4, col, item.value)
        for cell in (header, value):
            cell.fill = kpi_fill
            cell.font = white_bold
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value.number_format = _number_format(item.channel_unit)
        report.column_dimensions[get_column_letter(col)].width = 20

    # Plot images are embedded below statistics in a deterministic grid.
    chart_start_row = bottom_start_row + len(config.bottom_operations) + 3
    horizontal_span = max(8, math.ceil(channel_count / config.plot_columns))
    vertical_span = 20
    for plot_index, plot_id in enumerate(config.plot_ids):
        item = plots_by_id[plot_id]
        grid_row = plot_index // config.plot_columns
        grid_col = plot_index % config.plot_columns
        anchor_row = chart_start_row + grid_row * vertical_span
        anchor_col = 1 + grid_col * horizontal_span
        image = XLImage(item.output_file)
        image.width = config.plot_width_px
        image.height = config.plot_height_px
        report.add_image(image, f"{get_column_letter(anchor_col)}{anchor_row}")

    report.freeze_panes = "B5"
    report.row_dimensions[1].height = 28
    report.row_dimensions[2].height = 24
    report.row_dimensions[3].height = 44
    report.row_dimensions[4].height = 26
    for col_index in range(1, channel_count + 1):
        report.column_dimensions[get_column_letter(col_index)].width = 17

    _write_metadata_sheet(
        metadata,
        input_file=input_file,
        report_config_file=report_config_file,
        config=config,
        math_config_file=math_config_file,
        statistics_config_file=statistics_config_file,
        plotting_config_file=plotting_config_file,
        statistics_result=statistics_result,
        report_channels=report_channels,
        statistics_by_id=statistics_by_id,
        plots_by_id=plots_by_id,
        header_fill=metadata_fill,
        white_bold=white_bold,
        border=border,
    )

    workbook.save(output_path)


def _write_sergio_reference_workbook(
    output_path: Path,
    config: ExcelReportConfig,
    *,
    input_file: Path,
    report_config_file: Path,
    math_config_file: Path | None,
    statistics_config_file: Path,
    plotting_config_file: Path,
    statistics_result: StatisticsResult,
    report_channels: list[ChannelInfo],
    values_by_id: Mapping[str, Any],
    statistics_by_id: Mapping[str, StatisticResult],
    plots_by_id: Mapping[str, Any],
    native_chart_definitions: Mapping[str, Any],
) -> None:
    """Write the compact client-style layout reverse-engineered from Sergio's workbook.

    The visible report intentionally reserves rows 1-2 for RMS, rows 3-4 for
    channel/KPI headers, starts samples on row 5, writes configured per-channel
    summary values immediately after the final sample, and places plot assets
    under the KPI strip. Engineering calculations remain external to Excel.
    """

    workbook = Workbook()
    report = workbook.active
    report.title = config.report_sheet
    metadata = workbook.create_sheet(config.metadata_sheet)

    # Reference-inspired palette: raw channels stay white; calculated channels
    # use the same pale-yellow visual cue found in the client workbook.
    border_side = Side(style="thin", color="A6A6A6")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    raw_fill = PatternFill("solid", fgColor="FFFFFF")
    math_fill = PatternFill("solid", fgColor="FFF2CC")
    rms_fill = PatternFill("solid", fgColor="5B9BD5")
    metadata_fill = PatternFill("solid", fgColor="5B9BD5")
    white_bold = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
    header_font = Font(name="Calibri", size=11, color="000000", bold=False)

    channel_count = len(report_channels)
    channel_column_by_id = {
        channel.channel_id: index for index, channel in enumerate(report_channels, start=1)
    }

    # Rows 1 and 2: RMS labels and values directly above the relevant channel.
    for statistic_id in config.top_rms_statistic_ids:
        item = statistics_by_id[statistic_id]
        col = channel_column_by_id[item.channel_id]
        if statistic_id in config.rms_merges:
            report.merge_cells(config.rms_merges[statistic_id])
            col = report[config.rms_merges[statistic_id].split(":")[0]].column
        label = report.cell(1, col, item.display_name)
        value = report.cell(2, col, item.value)
        for cell in (label, value):
            cell.fill = rms_fill
            cell.font = white_bold
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value.number_format = _number_format(item.channel_unit)

        # Mirror the client helper-column convention when a squared helper is
        # immediately adjacent to the RMS source channel.
        if col < channel_count:
            helper = report_channels[col]
            if "squared" in helper.channel_id and item.channel_id in helper.dependencies:
                report.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)

    # Rows 3 and 4: selected channel name and unit, with math channels highlighted.
    for col_index, channel in enumerate(report_channels, start=1):
        fill = math_fill if channel.kind == "math" else raw_fill
        name_cell = report.cell(3, col_index, channel.display_name)
        unit_cell = report.cell(4, col_index, channel.unit or "")
        for cell in (name_cell, unit_cell):
            cell.fill = fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data begins on row 5, exactly like the reference workbook.
    data_start_row = 5
    sample_count = statistics_result.sample_count
    for col_index, channel in enumerate(report_channels, start=1):
        values = values_by_id[channel.channel_id]
        for row_offset, value in enumerate(values):
            cell = report.cell(data_start_row + row_offset, col_index, float(value))
            cell.number_format = _number_format(channel.unit)

    data_end_row = data_start_row + sample_count - 1

    # Per-channel bottom summary. First configured statistic for a channel goes
    # directly below the data; a second statistic (e.g. battery MIN) uses the
    # following row. This reproduces the MAX/MIN/last intent without a mixed
    # operation label row.
    bottom_offsets: dict[str, int] = {}
    for statistic_id in config.bottom_summary_statistic_ids:
        item = statistics_by_id[statistic_id]
        if item.channel_id not in channel_column_by_id:
            continue
        col = channel_column_by_id[item.channel_id]
        offset = bottom_offsets.get(item.channel_id, 0)
        row = data_end_row + 1 + offset
        bottom_offsets[item.channel_id] = offset + 1
        cell = report.cell(row, col, item.value)
        cell.number_format = _number_format(item.channel_unit)
        cell.font = Font(name="Calibri", size=11, bold=False)
        cell.border = Border(top=border_side)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # One blank separator column followed by the secondary KPI selection in rows 3-4.
    kpi_start_col = channel_count + config.blank_separator_columns + 1
    for index, statistic_id in enumerate(config.kpi_statistic_ids):
        item = statistics_by_id[statistic_id]
        col = kpi_start_col + index
        header = report.cell(3, col, item.display_name)
        value = report.cell(4, col, item.value)
        for cell in (header, value):
            cell.fill = raw_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        value.number_format = _number_format(item.channel_unit)
        report.column_dimensions[get_column_letter(col)].width = max(12, config.channel_width)

    # Native Excel charts sit directly below the KPI/secondary-selection strip,
    # as in the supplied report. The PNG path is retained for PowerPoint and for
    # legacy Excel configs that do not request native charts.
    if config.plot_placement == "kpi_panel":
        chart_start_row = 6
        chart_start_col = kpi_start_col
    else:
        max_bottom_depth = max(bottom_offsets.values(), default=0)
        chart_start_row = data_end_row + max_bottom_depth + 3
        chart_start_col = 1

    approx_col_px = 64
    approx_row_px = 20
    horizontal_span = max(8, math.ceil(config.plot_width_px / approx_col_px) + 1)
    vertical_span = max(16, math.ceil(config.plot_height_px / approx_row_px) + 2)
    native_chart_ids = [plot_id for plot_id in config.native_chart_ids if plot_id in native_chart_definitions]
    for plot_index, plot_id in enumerate(native_chart_ids):
        definition = native_chart_definitions[plot_id]
        grid_row = plot_index // config.plot_columns
        grid_col = plot_index % config.plot_columns
        anchor_row = chart_start_row + grid_row * vertical_span
        anchor_col = chart_start_col + grid_col * horizontal_span
        layout = config.chart_layout.get(plot_id)
        anchor = layout["anchor"] if layout else f"{get_column_letter(anchor_col)}{anchor_row}"
        chart = _build_native_scatter_chart(
            definition,
            report_sheet=report,
            channel_column_by_id=channel_column_by_id,
            data_start_row=data_start_row,
            data_end_row=data_end_row,
            chart_index=plot_index + 1,
        )
        chart.width = layout["width"] if layout else config.plot_width_px / 96
        chart.height = layout["height"] if layout else config.plot_height_px / 96
        report.add_chart(chart, anchor)

    for plot_index, plot_id in enumerate(config.plot_ids):
        item = plots_by_id[plot_id]
        grid_row = (plot_index + len(native_chart_ids)) // config.plot_columns
        grid_col = (plot_index + len(native_chart_ids)) % config.plot_columns
        anchor_row = chart_start_row + grid_row * vertical_span
        anchor_col = chart_start_col + grid_col * horizontal_span
        image = XLImage(item.output_file)
        image.width = config.plot_width_px
        image.height = config.plot_height_px
        report.add_image(image, f"{get_column_letter(anchor_col)}{anchor_row}")

    # Match the visible mechanics of the supplied workbook.
    report.freeze_panes = "B6"
    report.row_dimensions[2].height = 16
    report.row_dimensions[3].height = config.header_row_height
    report.row_dimensions[4].height = config.unit_row_height
    for col_index in range(1, channel_count + 1):
        report.column_dimensions[get_column_letter(col_index)].width = max(13, config.channel_width)
    for gap in range(channel_count + 1, kpi_start_col):
        report.column_dimensions[get_column_letter(gap)].width = 2.5

    # Give the plot/KPI panel stable widths so embedded images do not crowd the data area.
    if config.plot_ids or config.native_chart_ids:
        panel_end = 118 if config.native_chart_ids else chart_start_col + config.plot_columns * horizontal_span
        for col_index in range(chart_start_col, panel_end + 1):
            report.column_dimensions[get_column_letter(col_index)].width = 13

    _write_metadata_sheet(
        metadata,
        input_file=input_file,
        report_config_file=report_config_file,
        config=config,
        math_config_file=math_config_file,
        statistics_config_file=statistics_config_file,
        plotting_config_file=plotting_config_file,
        statistics_result=statistics_result,
        report_channels=report_channels,
        statistics_by_id=statistics_by_id,
        plots_by_id=plots_by_id,
        header_fill=metadata_fill,
        white_bold=white_bold,
        border=border,
    )

    workbook.save(output_path)


def _build_native_scatter_chart(
    definition: Any,
    *,
    report_sheet: Any,
    channel_column_by_id: Mapping[str, int],
    data_start_row: int,
    data_end_row: int,
    chart_index: int,
) -> ScatterChart:
    base_axis_id = 100000 + chart_index * 10
    chart = _make_scatter_shell(
        definition.title,
        x_title=definition.x_label or definition.x_channel_id,
        y_title=definition.primary_y_label or "Value",
        x_axis_id=base_axis_id + 1,
        y_axis_id=base_axis_id + 2,
        y_position="l",
    )
    x_col = channel_column_by_id[definition.x_channel_id]
    x_values = Reference(
        report_sheet,
        min_col=x_col,
        min_row=data_start_row,
        max_row=data_end_row,
    )
    secondary_chart = _make_scatter_shell(
        definition.title,
        x_title="",
        y_title=definition.secondary_y_label or "Value",
        x_axis_id=base_axis_id + 3,
        y_axis_id=base_axis_id + 4,
        y_position="r",
    )
    has_secondary = False
    for item in definition.series:
        y_col = channel_column_by_id[item.channel_id]
        y_values = Reference(
            report_sheet,
            min_col=y_col,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        series = Series(y_values, x_values, title=item.label or item.channel_id)
        series.graphicalProperties.line.width = 12700
        if item.axis == "secondary":
            has_secondary = True
            series.graphicalProperties.line.dashStyle = "dash"
            secondary_chart.series.append(series)
        else:
            chart.series.append(series)
    if has_secondary:
        secondary_chart.y_axis.crosses = "max"
        secondary_chart.x_axis.delete = True
        chart += secondary_chart
    return chart


def _make_scatter_shell(
    title: str,
    *,
    x_title: str,
    y_title: str,
    x_axis_id: int,
    y_axis_id: int,
    y_position: str,
) -> ScatterChart:
    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.scatterStyle = "line"
    chart.legend.position = "b"
    chart.x_axis.axId = x_axis_id
    chart.y_axis.axId = y_axis_id
    chart.x_axis.crossAx = y_axis_id
    chart.y_axis.crossAx = x_axis_id
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = y_position
    chart.x_axis.crosses = "autoZero"
    chart.y_axis.crosses = "autoZero"
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    return chart


def _write_metadata_sheet(
    sheet: Any,
    *,
    input_file: Path,
    report_config_file: Path,
    config: ExcelReportConfig,
    math_config_file: Path | None,
    statistics_config_file: Path,
    plotting_config_file: Path,
    statistics_result: StatisticsResult,
    report_channels: list[ChannelInfo],
    statistics_by_id: Mapping[str, StatisticResult],
    plots_by_id: Mapping[str, Any],
    header_fill: PatternFill,
    white_bold: Font,
    border: Border,
) -> None:
    sheet.freeze_panes = "A2"
    sheet.append(["VSM REPORT METADATA", "Value"])
    source_hash = statistics_result.dataset.quality.source_sha256
    metadata_rows = [
        ("Source file", client_display_filename(input_file)),
        ("Source SHA-256", source_hash),
        *_client_safe_pipeline_provenance(input_file),
        ("Samples", statistics_result.sample_count),
        ("Time channel", statistics_result.dataset.quality.time_channel_id or ""),
        ("Time start", statistics_result.dataset.quality.time_start),
        ("Time end", statistics_result.dataset.quality.time_end),
        ("Nominal time step", statistics_result.dataset.quality.nominal_time_step),
        ("Report configuration", Path(report_config_file).name),
        ("Math configuration", Path(math_config_file).name if math_config_file else ""),
        ("Statistics configuration", Path(statistics_config_file).name),
        ("Plotting configuration", Path(plotting_config_file).name),
    ]
    for row in metadata_rows:
        sheet.append(list(row))

    row = sheet.max_row + 2
    sheet.cell(row, 1, "REPORT CHANNELS")
    headers = ["Order", "channel_id", "display_name", "unit", "kind", "provenance", "dependencies"]
    for col, value in enumerate(headers, start=1):
        sheet.cell(row + 1, col, value)
    for index, channel in enumerate(report_channels, start=1):
        values = [
            index,
            channel.channel_id,
            channel.display_name,
            channel.unit or "",
            channel.kind,
            channel.provenance.replace(input_file.name, client_display_filename(input_file)),
            ";".join(channel.dependencies),
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row + 1 + index, col, value)

    row = sheet.max_row + 2
    sheet.cell(row, 1, "STATISTICS")
    stat_headers = ["statistic_id", "channel_id", "operation", "placement_group", "value", "unit", "display_name"]
    for col, value in enumerate(stat_headers, start=1):
        sheet.cell(row + 1, col, value)
    for index, item in enumerate(statistics_by_id.values(), start=1):
        values = [item.statistic_id, item.channel_id, item.operation, item.placement_group, item.value, item.channel_unit or "", item.display_name]
        for col, value in enumerate(values, start=1):
            sheet.cell(row + 1 + index, col, value)

    row = sheet.max_row + 2
    sheet.cell(row, 1, "PLOTS")
    plot_headers = ["plot_id", "title", "x_channel_id", "primary_series", "secondary_series", "output_file"]
    for col, value in enumerate(plot_headers, start=1):
        sheet.cell(row + 1, col, value)
    for index, item in enumerate(plots_by_id.values(), start=1):
        values = [
            item.plot_id,
            item.title,
            item.x_channel_id,
            ";".join(item.primary_series_ids),
            ";".join(item.secondary_series_ids),
            Path(item.output_file).name,
        ]
        for col, value in enumerate(values, start=1):
            sheet.cell(row + 1 + index, col, value)

    # Style section labels and table headers.
    for row_index in range(1, sheet.max_row + 1):
        first = sheet.cell(row_index, 1).value
        if first in {"VSM REPORT METADATA", "REPORT CHANNELS", "STATISTICS", "PLOTS"}:
            for col in range(1, 8):
                cell = sheet.cell(row_index, col)
                cell.fill = header_fill
                cell.font = white_bold
                cell.border = border
        if first in {"Order", "statistic_id", "plot_id"}:
            for col in range(1, 8):
                cell = sheet.cell(row_index, col)
                cell.fill = header_fill
                cell.font = white_bold
                cell.border = border

    for col, width in {"A": 28, "B": 42, "C": 30, "D": 18, "E": 18, "F": 45, "G": 45}.items():
        sheet.column_dimensions[col].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _client_safe_pipeline_provenance(input_file: Path) -> list[tuple[str, object]]:
    if input_file.name != "duty_cycle_dataset.csv":
        return []
    root = input_file.parent.parent
    rows: list[tuple[str, object]] = [("Internal pipeline artifact", input_file.name)]
    inspection_path = root / "01_inspection" / "inspection_result.json"
    if inspection_path.exists():
        try:
            inspection = json.loads(inspection_path.read_text(encoding="utf-8"))
            source_path = Path(str(inspection.get("source_path", "")))
            quality = inspection.get("quality", {}) if isinstance(inspection.get("quality"), dict) else {}
            if source_path.name:
                rows.append(("Original VSM source workbook", source_path.name))
            if quality.get("source_sha256"):
                rows.append(("Original VSM source SHA-256", quality["source_sha256"]))
        except (OSError, json.JSONDecodeError):
            pass
    summary_path = input_file.parent / "duty_cycle_summary.txt"
    if summary_path.exists():
        try:
            for line in summary_path.read_text(encoding="utf-8").splitlines():
                if line.startswith("Scenario: "):
                    rows.append(("Duty-cycle scenario ID", line.split(": ", 1)[1]))
                    break
        except OSError:
            pass
    profile_path = input_file.parent / "profile_provenance.csv"
    if profile_path.exists():
        try:
            with profile_path.open(encoding="utf-8-sig", newline="") as handle:
                first = next(csv.DictReader(handle), None)
            if first:
                rows.append(("External profile provider ID", first.get("provider_id", "")))
                rows.append(("External profile/reference workbook", first.get("source_file", "")))
                rows.append(("External profile/reference SHA-256", first.get("source_sha256", "")))
        except (OSError, csv.Error, StopIteration):
            pass
    return [(label, value) for label, value in rows if value not in (None, "")]


def _manifest(
    result: ExcelReportResult,
    statistics_config_file: str | Path,
    plotting_config_file: str | Path,
    math_config_file: str | Path | None,
) -> dict[str, Any]:
    payload = {
        "configuration_version": result.config.version,
        "software_version": __version__,
        "source_file": str(result.statistics_result.dataset.source_path),
        "source_sha256": result.statistics_result.dataset.quality.source_sha256,
        "report_configuration_file": str(result.config_path),
        "report_configuration_sha256": sha256_file(result.config_path),
        "statistics_configuration_file": str(Path(statistics_config_file).expanduser().resolve()),
        "statistics_configuration_sha256": sha256_file(Path(statistics_config_file).expanduser().resolve()),
        "plotting_configuration_file": str(Path(plotting_config_file).expanduser().resolve()),
        "plotting_configuration_sha256": sha256_file(Path(plotting_config_file).expanduser().resolve()),
        "sample_count": result.sample_count,
        "report_channel_count": result.channel_count,
        "statistic_count": result.statistic_count,
        "configured_plot_count": result.configured_plot_count,
        "plot_count": result.configured_plot_count,
        "plot_series_count": result.plotting_result.series_count,
        "native_excel_chart_count": result.native_excel_chart_count,
        "embedded_plot_image_count": result.embedded_plot_image_count,
        "report_channel_ids": [channel.channel_id for channel in result.report_channels],
        "top_rms_statistic_ids": list(result.config.top_rms_statistic_ids),
        "kpi_statistic_ids": list(result.config.kpi_statistic_ids),
        "bottom_operations": list(result.config.bottom_operations),
        "bottom_summary_statistic_ids": list(result.config.bottom_summary_statistic_ids),
        "layout_profile": result.config.layout_profile,
        "plot_placement": result.config.plot_placement,
        "plot_ids": list(result.config.plot_ids),
        "report_file": str(result.report_path),
    }
    if math_config_file is not None:
        path = Path(math_config_file).expanduser().resolve()
        payload["math_configuration_file"] = str(path)
        payload["math_configuration_sha256"] = sha256_file(path)
    else:
        payload["math_configuration_file"] = None
        payload["math_configuration_sha256"] = None
    return payload


def _summary(result: ExcelReportResult) -> str:
    lines = [
        "VSM EXCEL REPORT SUMMARY",
        "========================",
        f"Status: PASS",
        f"Source: {result.statistics_result.dataset.source_path}",
        f"Samples: {result.sample_count}",
        f"Report channels: {result.channel_count}",
        f"Statistics available: {result.statistic_count}",
        f"Native Excel charts embedded: {result.native_excel_chart_count}",
        f"Plot images embedded: {result.embedded_plot_image_count}",
        f"Configured plots rendered: {result.configured_plot_count}",
        f"Plot series rendered: {result.plotting_result.series_count}",
        f"Layout profile: {result.config.layout_profile}",
        f"Plot placement: {result.config.plot_placement}",
        f"Workbook: {result.report_path}",
        "",
        "Top RMS:",
    ]
    stats_by_id = {item.statistic_id: item for item in result.statistics_result.statistics}
    for statistic_id in result.config.top_rms_statistic_ids:
        item = stats_by_id[statistic_id]
        lines.append(f"  - {item.display_name}: {item.value:.12g} [{item.channel_unit or '-'}]")
    lines.append("")
    lines.append("KPI strip:")
    for statistic_id in result.config.kpi_statistic_ids:
        item = stats_by_id[statistic_id]
        lines.append(f"  - {item.display_name}: {item.value:.12g} [{item.channel_unit or '-'}]")
    return "\n".join(lines) + "\n"


def _number_format(unit: str | None) -> str:
    if unit is None:
        return "0.000000"
    normalized = unit.strip().lower()
    if normalized in {"%", "percent"}:
        return "0.0000"
    if normalized in {"s", "min", "m", "km", "kph", "rpm", "nm", "kw", "kwh", "kg"}:
        return "0.000000"
    return "0.000000"


def _load_channel_metadata(raw: object, channel_ids: tuple[str, ...]) -> dict[str, dict[str, str]]:
    if raw is None:
        return {}
    if not isinstance(raw, dict):
        raise ConfigurationError("channel_metadata must be a YAML mapping")
    allowed = {"display_name", "unit", "kind"}
    metadata: dict[str, dict[str, str]] = {}
    for channel_id, item in raw.items():
        if not isinstance(channel_id, str) or not channel_id.strip():
            raise ConfigurationError("channel_metadata keys must be non-empty channel IDs")
        if channel_id not in channel_ids:
            raise ConfigurationError(f"channel_metadata contains channel not listed in channels: {channel_id}")
        if not isinstance(item, dict):
            raise ConfigurationError(f"channel_metadata.{channel_id} must be a YAML mapping")
        _reject_unknown_keys(item, allowed, f"channel_metadata.{channel_id}")
        override: dict[str, str] = {}
        for key in allowed:
            value = item.get(key)
            if value is None:
                continue
            override[key] = _nonempty_string(value, f"channel_metadata.{channel_id}.{key}")
        if "kind" in override and override["kind"] not in {"raw", "math"}:
            raise ConfigurationError(f"channel_metadata.{channel_id}.kind must be raw or math")
        metadata[channel_id] = override
    return metadata


def _load_chart_layout(raw: object) -> dict[str, dict[str, Any]]:
    if raw is None:
        return {}
    if not isinstance(raw, dict):
        raise ConfigurationError("plots.chart_layout must be a YAML mapping")
    layouts: dict[str, dict[str, Any]] = {}
    for plot_id, item in raw.items():
        if not isinstance(plot_id, str) or not plot_id.strip():
            raise ConfigurationError("plots.chart_layout keys must be non-empty plot IDs")
        if not isinstance(item, dict):
            raise ConfigurationError(f"plots.chart_layout.{plot_id} must be a YAML mapping")
        _reject_unknown_keys(item, {"anchor", "width", "height"}, f"plots.chart_layout.{plot_id}")
        layouts[plot_id] = {
            "anchor": _nonempty_string(item.get("anchor"), f"plots.chart_layout.{plot_id}.anchor").upper(),
            "width": _positive_number(item.get("width", 15.0), f"plots.chart_layout.{plot_id}.width", minimum=6.0, maximum=30.0),
            "height": _positive_number(item.get("height", 7.5), f"plots.chart_layout.{plot_id}.height", minimum=4.0, maximum=16.0),
        }
    return layouts


def _load_string_mapping(raw: object, context: str) -> dict[str, str]:
    if raw is None:
        return {}
    if not isinstance(raw, dict):
        raise ConfigurationError(f"{context} must be a YAML mapping")
    return {
        _nonempty_string(key, f"{context} key"): _nonempty_string(value, f"{context}.{key}")
        for key, value in raw.items()
    }


def _apply_channel_metadata(channel: ChannelInfo, metadata: Mapping[str, str] | None) -> ChannelInfo:
    if not metadata:
        return channel
    return replace(
        channel,
        display_name=metadata.get("display_name", channel.display_name),
        unit=metadata.get("unit", channel.unit or None),
        kind=metadata.get("kind", channel.kind),
    )


def _plain_xlsx_filename(value: object, context: str) -> str:
    filename = _nonempty_string(value, context)
    path = Path(filename)
    if path.name != filename or path.suffix.lower() != ".xlsx":
        raise ConfigurationError(f"{context} must be a plain .xlsx filename without directories")
    return filename


def _sheet_name(value: object, context: str) -> str:
    name = _nonempty_string(value, context)
    if len(name) > 31 or any(character in name for character in "[]:*?/\\"):
        raise ConfigurationError(f"{context} is not a valid Excel sheet name")
    return name


def _string_list(value: object, context: str) -> tuple[str, ...]:
    if not isinstance(value, list):
        raise ConfigurationError(f"{context} must be a YAML list")
    items = tuple(_nonempty_string(item, f"{context}[]") for item in value)
    duplicates = _duplicates(items)
    if duplicates:
        raise ConfigurationError(f"{context} contains duplicates: " + ", ".join(duplicates))
    return items


def _nonempty_string(value: object, context: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ConfigurationError(f"{context} must be a non-empty string")
    return value.strip()


def _optional_string(value: object, context: str) -> str | None:
    if value is None:
        return None
    return _nonempty_string(value, context)


def _choice(value: object, context: str, allowed: set[str]) -> str:
    choice = _nonempty_string(value, context)
    if choice not in allowed:
        raise ConfigurationError(f"{context} must be one of: " + ", ".join(sorted(allowed)))
    return choice


def _positive_int(value: object, context: str, minimum: int = 1, maximum: int | None = None) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < minimum:
        raise ConfigurationError(f"{context} must be an integer >= {minimum}")
    if maximum is not None and value > maximum:
        raise ConfigurationError(f"{context} must be <= {maximum}")
    return value


def _positive_number(value: object, context: str, minimum: float = 0.0, maximum: float | None = None) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ConfigurationError(f"{context} must be a positive finite number")
    result = float(value)
    if not math.isfinite(result) or result < minimum:
        raise ConfigurationError(f"{context} must be >= {minimum}")
    if maximum is not None and result > maximum:
        raise ConfigurationError(f"{context} must be <= {maximum}")
    return result


def _reject_unknown_keys(mapping: Mapping[str, object], allowed: set[str], context: str) -> None:
    unknown = sorted(set(mapping) - allowed)
    if unknown:
        raise ConfigurationError(f"Unknown key(s) in {context}: " + ", ".join(unknown))


def _duplicates(values: tuple[str, ...]) -> list[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for value in values:
        if value in seen:
            duplicates.add(value)
        seen.add(value)
    return sorted(duplicates)
